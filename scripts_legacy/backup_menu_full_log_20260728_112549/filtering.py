from __future__ import annotations

import csv
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .config import AppConfig


def detect_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "gb18030", "gbk", "cp936", "utf-16"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                file.read(8192)
            return encoding
        except UnicodeError:
            continue
    return "gb18030"


def read_table(path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows_iter = worksheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows_iter)]
            rows = []
            for row in rows_iter:
                rows.append({headers[index]: str(value or "") for index, value in enumerate(row[: len(headers)])})
            return headers, rows, "excel"
        finally:
            workbook.close()
    encoding = detect_encoding(path)
    with path.open("r", encoding=encoding, newline="") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        rows = [{key: (value or "") for key, value in row.items()} for row in reader]
    return headers, rows, encoding


def normalize(value: str) -> str:
    return value.replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")


def variants(name: str) -> list[str]:
    names = [name]
    if "_" in name:
        names.append(name.replace("_", "*"))
    if "*" in name:
        names.append(name.replace("*", "_"))
    return [item for item in dict.fromkeys(names) if item]


def find_column(headers: list[str], candidates: list[str], required: bool = True) -> str | None:
    expanded = []
    for candidate in candidates:
        expanded.extend(variants(str(candidate or "")))
    expanded = [item for item in expanded if item]
    normalized = {normalize(header): header for header in headers}
    for candidate in expanded:
        if normalize(candidate) in normalized:
            return normalized[normalize(candidate)]
    for header in headers:
        compact = normalize(header)
        if any(normalize(candidate) in compact for candidate in expanded):
            return header
    if required:
        raise KeyError("缺少必要字段：" + "、".join(expanded))
    return None


def to_decimal(value: object) -> Decimal:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def format_decimal(value: object) -> str:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return str(int(value))
        text = format(value.normalize(), "f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    return str(value or "")


def add_audit(rows: list[dict[str, str]], item: str, status: str, detail: str) -> None:
    rows.append({"检查项": item, "状态": status, "说明": detail})


def audit_status(rows: list[dict[str, str]]) -> str:
    statuses = {row["状态"] for row in rows}
    if "失败" in statuses:
        return "FAIL"
    if "预警" in statuses:
        return "WARN"
    return "PASS"


def load_sku_set(path_text: str, sku_column_name: str) -> set[str]:
    if not path_text:
        return set()
    path = Path(path_text)
    if not path.exists():
        raise FileNotFoundError(f"SKU名单文件不存在：{path}")
    headers, rows, _ = read_table(path)
    sku_column = find_column(headers, [sku_column_name, "sku", "SKU", "SKU(含影)"], required=False) or headers[0]
    return {str(row.get(sku_column, "")).strip() for row in rows if str(row.get(sku_column, "")).strip()}


def dedupe(rows: list[dict[str, str]], sku_column: str, metric_columns: list[str]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        sku = str(row.get(sku_column, "")).strip()
        if sku not in grouped:
            grouped[sku] = dict(row)
            for column in metric_columns:
                grouped[sku][column] = to_decimal(row.get(column, ""))
        else:
            for column in metric_columns:
                grouped[sku][column] = to_decimal(grouped[sku].get(column, "")) + to_decimal(row.get(column, ""))
    return list(grouped.values())


def remove_skus(rows: list[dict[str, Any]], sku_column: str, skus: set[str]) -> tuple[list[dict[str, Any]], int]:
    if not skus:
        return rows, 0
    kept = [row for row in rows if str(row.get(sku_column, "")).strip() not in skus]
    return kept, len(rows) - len(kept)


def write_sku_file(path: Path, skus: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "汇总"
    worksheet.append(["sku"])
    for sku in skus:
        worksheet.append([sku])
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 22
    workbook.save(path)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: format_decimal(row.get(header, "")) for header in headers})



def write_result_xlsx(path: Path, headers: list[str], rows: list[dict[str, Any]], logs: list[str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "筛选结果"
    sheet.append(headers)
    for row in rows:
        sheet.append([format_decimal(row.get(header, "")) for header in headers])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(column[0].value or "")) + 2, 12), 28)

    log_sheet = workbook.create_sheet("筛品日志")
    for line in logs:
        log_sheet.append([line])
    log_sheet.column_dimensions["A"].width = 120
    for row in log_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(path)

def validate_columns(headers: list[str], config: AppConfig) -> tuple[dict[str, str], list[str]]:
    specs = {
        "sku": [config.get("sku_column"), "SKU(含影)", "sku", "SKU"],
        "name": [config.get("product_name_column"), "sku名称", "商品名称"],
        "mapping": [config.get("mapping_column"), "1自营3POP_映射"],
        "shop": [config.get("shop_column"), "店铺名称"],
        "department": [config.get("department_column"), "1级部门名称（类目运营）", "一级部门"],
        "category2": [config.get("category2_column"), "商品二级分类名称"],
        "erp": [config.get("erp_column"), "销售员ERP帐号", "销售员ERP账号"],
    }
    found: dict[str, str] = {}
    missing: list[str] = []
    for key, candidates in specs.items():
        column = find_column(headers, [str(item) for item in candidates], required=False)
        if column:
            found[key] = column
        else:
            missing.append("/".join([str(item) for item in candidates if item]))
    metric_columns = []
    for metric in config.get("metric_sum_columns", []):
        column = find_column(headers, [str(metric)], required=False)
        if column:
            metric_columns.append(column)
        else:
            missing.append(str(metric))
    found["metrics"] = "||".join(metric_columns)
    return found, missing


def run_filtering(config: AppConfig) -> dict[str, Any]:
    source = Path(str(config.get("source_file")))
    output_dir = config.filter_output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    date_tag = str(config.get("date_tag"))
    audit: list[dict[str, str]] = []

    if not source.exists():
        add_audit(audit, "源文件存在", "失败", f"源文件不存在：{source}")
        raise FileNotFoundError(f"源文件不存在：{source}")
    add_audit(audit, "源文件存在", "通过", f"源文件路径：{source}")

    headers, source_rows, source_type = read_table(source)
    add_audit(audit, "源文件读取", "通过", f"已读取 {len(source_rows)} 条，格式：{source_type}")
    columns, missing = validate_columns(headers, config)
    if missing:
        add_audit(audit, "必填字段完整", "失败", "缺少字段：" + "、".join(missing))
        fail_report = output_dir / f"筛品失败审核报告_{date_tag}.txt"
        fail_report.write_text("\n".join([f"{row['检查项']}｜{row['状态']}｜{row['说明']}" for row in audit]), encoding="utf-8")
        raise KeyError(f"关键字段缺失，已终止：{', '.join(missing)}；失败报告：{fail_report}")
    add_audit(audit, "必填字段完整", "通过", "所有关键字段存在")

    sku_col = columns["sku"]
    name_col = columns["name"]
    mapping_col = columns["mapping"]
    shop_col = columns["shop"]
    dept_col = columns["department"]
    category_col = columns["category2"]
    erp_col = columns["erp"]
    metric_cols = columns["metrics"].split("||")

    logs: list[str] = []
    targets = list(config.get("target_departments", []))
    step1 = [row for row in source_rows if any(keyword in str(row.get(dept_col, "")) for keyword in targets)]
    logs.append(f"1、取数模型提取近30天动销>200，一级部门为“{'、'.join(targets)}”的sku获得{len(step1)}个sku记录（未去重）。")

    keep_values = set(config.get("keep_mapping_values", []))
    pop_keyword = str(config.get("keep_pop_shop_keyword", ""))
    keep_rows = [row for row in step1 if str(row.get(mapping_col, "")).strip() in keep_values]
    pop_rows = [row for row in step1 if str(row.get(mapping_col, "")).strip().upper() == "POP" and pop_keyword in str(row.get(shop_col, ""))]
    step2 = [row for row in step1 if str(row.get(mapping_col, "")).strip() in keep_values or (str(row.get(mapping_col, "")).strip().upper() == "POP" and pop_keyword in str(row.get(shop_col, "")))]
    step2_removed = len(step1) - len(step2)
    logs.append(f"2、保留{'+'.join(keep_values)}（{len(keep_rows)}）、{pop_keyword}（{len(pop_rows)}），剔除{step2_removed}个，剩余{len(step2)}个。")
    if step1 and step2_removed / len(step1) > 0.6:
        add_audit(audit, "第2步剔除比例", "预警", f"第2步剔除比例 {step2_removed / len(step1):.1%}，请确认自营/京喜规则")
    else:
        add_audit(audit, "第2步剔除比例", "通过", f"第2步剔除 {step2_removed} 个")

    excluded_categories = set(config.get("exclude_category2", []))
    counts = Counter(str(row.get(category_col, "")).strip() for row in step2)
    step3 = [row for row in step2 if str(row.get(category_col, "")).strip() not in excluded_categories]
    category_text = "、".join([f"“{category}”（{counts[category]}个）" for category in config.get("exclude_category2", [])])
    logs.append(f"3、剔除{category_text}，剩余{len(step3)}个。")

    step4 = dedupe(step3, sku_col, metric_cols)
    logs.append(f"4、sku去重后{len(step4)}个，并对“{'、'.join(metric_cols)}”按sku求和。")
    add_audit(audit, "SKU去重检查", "通过", f"去重前 {len(step3)}，去重后 {len(step4)}")

    keywords = list(config.get("exclude_name_keywords", []))
    keyword_counts = {keyword: sum(1 for row in step4 if keyword in str(row.get(name_col, ""))) for keyword in keywords}
    step5 = [row for row in step4 if not any(keyword in str(row.get(name_col, "")) for keyword in keywords)]
    keyword_text = "、".join([f"“{keyword}”（{count}个）" for keyword, count in keyword_counts.items()])
    logs.append(f"5、删除{keyword_text}，实际剔除{len(step4) - len(step5)}个，剩余{len(step5)}个。")

    current_success_file = str(config.get("current_success_skus_file", "") or "").strip()
    if current_success_file:
        step6, removed6 = remove_skus(step5, sku_col, load_sku_set(current_success_file, sku_col))
        logs.append(f"6、删除目前提报成功的sku {removed6}个，剩余{len(step6)}个。")
        add_audit(audit, "目前提报成功SKU名单", "通过", f"已读取并剔除 {removed6} 个")
    else:
        step6 = step5
        logs.append(f"6、未执行：未提供“目前提报成功sku”名单，保留剩余{len(step6)}个记录。")
        add_audit(audit, "目前提报成功SKU名单", "预警", "未提供名单，步骤6未执行")

    submitted_file = str(config.get("submitted_skus_file", "") or "").strip()
    if submitted_file:
        step8, removed8 = remove_skus(step6, sku_col, load_sku_set(submitted_file, sku_col))
        logs.append(f"7、剔除已经提报过的sku {removed8}个，剩余{len(step8)}个。")
        add_audit(audit, "已经提报SKU名单", "通过", f"已读取并剔除 {removed8} 个")
    else:
        step8 = step6
        logs.append(f"7、未执行：未提供“已经提报过sku”名单，保留剩余{len(step8)}个记录。")
        add_audit(audit, "已经提报SKU名单", "预警", "未提供名单，步骤8未执行；后续可在 config.yaml 填写 submitted_skus_file")

    excluded_erps = set(config.get("exclude_erp_accounts", []))
    removed_erp = [row for row in step8 if str(row.get(erp_col, "")).strip() in excluded_erps]
    final_rows = [row for row in step8 if str(row.get(erp_col, "")).strip() not in excluded_erps]
    logs.append(f"8、剔除指定ERP帐号（{'、'.join(excluded_erps)}）{len(removed_erp)}个，剩余{len(final_rows)}个。")
    add_audit(audit, "最终结果数量", "通过" if final_rows else "失败", f"最终剩余 {len(final_rows)} 个")

    skus = [str(row.get(sku_col, "")).strip() for row in final_rows if str(row.get(sku_col, "")).strip()]
    sku_file = output_dir / f"筛品结果SKU_{date_tag}.xlsx"
    write_sku_file(sku_file, skus)
    add_audit(audit, "下一模块SKU文件", "通过" if len(skus) == len(final_rows) else "预警", f"已生成 {len(skus)} 条：{sku_file}")

    xlsx = output_dir / f"新人价筛品结果_{date_tag}.xlsx"
    log_path = output_dir / f"新人价筛品日志_{date_tag}.txt"
    add_audit(audit, "输出文件检查", "通过", "已生成Excel、日志和SKU文件")
    write_result_xlsx(xlsx, headers, final_rows, logs)
    log_path.write_text("\n".join(logs) + "\n", encoding="utf-8")
    audit_report = output_dir / f"筛品审核报告_{date_tag}.txt"
    audit_lines = ["检查项\t状态\t说明"]
    audit_lines.extend([f"{row['检查项']}\t{row['状态']}\t{row['说明']}" for row in audit])
    audit_report.write_text("\n".join(audit_lines) + "\n", encoding="utf-8")

    csv_path = None
    if bool(config.get("export_csv", False)):
        csv_path = output_dir / f"新人价筛品结果_{date_tag}.csv"
        write_csv(csv_path, headers, final_rows)

    return {
        "status": audit_status(audit),
        "warnings": [row for row in audit if row["状态"] == "预警"],
        "failures": [row for row in audit if row["状态"] == "失败"],
        "source_file": source,
        "source_rows": len(source_rows),
        "rows": len(final_rows),
        "xlsx": xlsx,
        "log": log_path,
        "sku_file": sku_file,
        "audit_report": audit_report,
        "csv": csv_path,
        "logs": logs,
        "audit_rows": audit,
        "source_type": source_type,
    }
