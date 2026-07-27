from __future__ import annotations

import argparse
import csv
import sys
from collections import Counter
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path


DEFAULT_METRIC_COLUMNS = [
    "去重计数_用户PIN加密",
    "求和_成交金额",
    "去重计数_销售订单编号",
    "求和_销售数量",
]


@dataclass
class StepState:
    rows: list[dict[str, str | Decimal]]
    logs: list[str]


def detect_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "gb18030", "gbk", "cp936", "utf-16"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                file.read(8192)
            return encoding
        except UnicodeError:
            continue
    return "gb18030"


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    encoding = detect_encoding(path)
    with path.open("r", encoding=encoding, newline="") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        rows = [{key: (value or "") for key, value in row.items()} for row in reader]
    return headers, rows, encoding


def normalize_header(value: str) -> str:
    return value.replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")


def find_column(headers: list[str], candidates: list[str], required: bool = True) -> str | None:
    candidates = [candidate for candidate in candidates if candidate]
    if not candidates:
        if required:
            raise KeyError("缺少必要字段：未提供候选字段名")
        return None
    normalized = {normalize_header(header): header for header in headers}
    for candidate in candidates:
        if normalize_header(candidate) in normalized:
            return normalized[normalize_header(candidate)]
    for header in headers:
        compact = normalize_header(header)
        if any(normalize_header(candidate) in compact for candidate in candidates):
            return header
    if required:
        raise KeyError(f"缺少必要字段：{', '.join(candidates)}")
    return None

def to_decimal(value: object) -> Decimal:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def format_decimal(value: object) -> str:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return str(int(value))
        text = format(value.normalize(), "f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    return str(value or "")


def load_sku_set(path: Path | None, sku_column_name: str | None = None) -> set[str]:
    if not path:
        return set()
    headers, rows, _ = read_csv(path)
    if not headers:
        return set()
    sku_column = sku_column_name if sku_column_name in headers else find_column(headers, [sku_column_name or "", "SKU", "sku", "SKU(含影)"], required=False)
    if not sku_column:
        sku_column = headers[0]
    return {str(row.get(sku_column, "")).strip() for row in rows if str(row.get(sku_column, "")).strip()}


def write_csv(path: Path, headers: list[str], rows: list[dict[str, str | Decimal]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: format_decimal(row.get(header, "")) for header in headers})


def write_xlsx(path: Path, headers: list[str], rows: list[dict[str, str | Decimal]], log_lines: list[str]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except Exception:
        return False

    workbook = Workbook(write_only=False)
    result_sheet = workbook.active
    result_sheet.title = "筛选结果"
    result_sheet.append(headers)
    for row in rows:
        result_sheet.append([format_decimal(row.get(header, "")) for header in headers])
    for cell in result_sheet[1]:
        cell.font = Font(bold=True)
    result_sheet.freeze_panes = "A2"
    for column in result_sheet.columns:
        header = str(column[0].value or "")
        width = min(max(len(header) + 2, 12), 28)
        result_sheet.column_dimensions[column[0].column_letter].width = width

    log_sheet = workbook.create_sheet("筛品日志")
    for line in log_lines:
        log_sheet.append([line])
    log_sheet.column_dimensions["A"].width = 120
    for row in log_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)
    return True


def deduplicate_by_sku(rows: list[dict[str, str]], sku_column: str, metric_columns: list[str]) -> list[dict[str, str | Decimal]]:
    grouped: dict[str, dict[str, str | Decimal]] = {}
    for row in rows:
        sku = str(row.get(sku_column, "")).strip()
        if sku not in grouped:
            grouped[sku] = dict(row)
            for metric_column in metric_columns:
                grouped[sku][metric_column] = to_decimal(row.get(metric_column, ""))
        else:
            for metric_column in metric_columns:
                grouped[sku][metric_column] = to_decimal(grouped[sku].get(metric_column, "")) + to_decimal(row.get(metric_column, ""))
    return list(grouped.values())


def remove_skus(rows: list[dict[str, str | Decimal]], sku_column: str, skus: set[str]) -> tuple[list[dict[str, str | Decimal]], int]:
    if not skus:
        return rows, 0
    kept = [row for row in rows if str(row.get(sku_column, "")).strip() not in skus]
    return kept, len(rows) - len(kept)


def main() -> int:
    parser = argparse.ArgumentParser(description="新人价筛品自动化")
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--current-success-skus", type=Path)
    parser.add_argument("--submitted-skus", type=Path)
    parser.add_argument("--price-column")
    parser.add_argument("--price-threshold", type=Decimal, default=Decimal("200"))
    parser.add_argument("--date-tag", default="20260727")
    args = parser.parse_args()

    args.output_dir.mkdir(parents=True, exist_ok=True)
    headers, source_rows, encoding = read_csv(args.source)

    sku_column = find_column(headers, ["SKU(含影)", "SKU", "sku"])
    product_name_column = find_column(headers, ["sku名称", "商品名称"])
    map_column = find_column(headers, ["1自营3POP_映射"])
    shop_column = find_column(headers, ["店铺名称"])
    department_column = find_column(headers, ["1级部门名称（类目运营）", "一级部门"])
    erp_column = find_column(headers, ["销售员ERP帐号", "销售员ERP账号"])
    category2_column = find_column(headers, ["商品二级分类名称"])
    metric_columns = [column for column in DEFAULT_METRIC_COLUMNS if column in headers]
    missing_metric_columns = [column for column in DEFAULT_METRIC_COLUMNS if column not in headers]
    if missing_metric_columns:
        raise KeyError(f"缺少汇总字段：{', '.join(missing_metric_columns)}")

    target_departments = ("医药", "营养保健", "医疗器械", "消费器械")
    step1_rows = [row for row in source_rows if any(keyword in str(row.get(department_column, "")) for keyword in target_departments)]
    logs: list[str] = []
    logs.append(f"源文件：{args.source}")
    logs.append(f"读取编码：{encoding}")
    logs.append(f"1、取数模型提取近30天动销>200，一级部门为“医药、营养保健、医疗器械、消费器械”的sku获得{len(step1_rows)}个sku记录（未去重）。")
    if len(step1_rows) != len(source_rows):
        logs.append(f"   其中源文件总记录{len(source_rows)}个，因一级部门不在目标范围剔除{len(source_rows) - len(step1_rows)}个。")

    self_rows = [row for row in step1_rows if str(row.get(map_column, "")).strip() == "自营"]
    jingxi_rows = [row for row in step1_rows if str(row.get(map_column, "")).strip().upper() == "POP" and "京喜" in str(row.get(shop_column, ""))]
    step2_rows = [row for row in step1_rows if str(row.get(map_column, "")).strip() == "自营" or (str(row.get(map_column, "")).strip().upper() == "POP" and "京喜" in str(row.get(shop_column, "")))]
    logs.append(f"2、保留自营（{len(self_rows)}）、京喜（{len(jingxi_rows)}），剔除{len(step1_rows) - len(step2_rows)}个，剩余{len(step2_rows)}个。")

    excluded_categories = ["健康服务套餐", "其他服务", "特殊商品"]
    category_counts = Counter(str(row.get(category2_column, "")).strip() for row in step2_rows)
    excluded_total = sum(category_counts[category] for category in excluded_categories)
    step3_rows = [row for row in step2_rows if str(row.get(category2_column, "")).strip() not in excluded_categories]
    logs.append(f"3、剔除“健康服务套餐”（{category_counts['健康服务套餐']}个）、“其他服务”（{category_counts['其他服务']}个）、“特殊商品”（{category_counts['特殊商品']}个），剩余{len(step3_rows)}个。")
    if excluded_total != len(step2_rows) - len(step3_rows):
        logs.append("   分类剔除存在重叠或空值异常，请复核。")

    step4_rows = deduplicate_by_sku(step3_rows, sku_column, metric_columns)
    logs.append(f"4、按照sku去重，并对“去重计数_用户PIN加密、求和_成交金额、去重计数_销售订单编号、求和_销售数量”按sku求和，sku去重后{len(step4_rows)}个。")

    non_sale_rows = [row for row in step4_rows if "非卖品" in str(row.get(product_name_column, ""))]
    gift_rows = [row for row in step4_rows if "赠品" in str(row.get(product_name_column, ""))]
    both_keyword_rows = [row for row in step4_rows if "非卖品" in str(row.get(product_name_column, "")) and "赠品" in str(row.get(product_name_column, ""))]
    step5_rows = [row for row in step4_rows if "非卖品" not in str(row.get(product_name_column, "")) and "赠品" not in str(row.get(product_name_column, ""))]
    logs.append(f"5、删除“非卖品”（{len(non_sale_rows)}个）、“赠品”（{len(gift_rows)}个），其中同时命中{len(both_keyword_rows)}个，实际剔除{len(step4_rows) - len(step5_rows)}个，剩余{len(step5_rows)}个。")

    current_success_skus = load_sku_set(args.current_success_skus, sku_column) if args.current_success_skus else set()
    step6_rows, step6_removed = remove_skus(step5_rows, sku_column, current_success_skus)
    if args.current_success_skus:
        logs.append(f"6、删除目前提报成功的sku {step6_removed}个，剩余{len(step6_rows)}个。")
    else:
        logs.append(f"6、未执行：未提供“目前提报成功sku”名单，继续保留上一步{len(step6_rows)}个。")

    price_column = args.price_column if args.price_column in headers else find_column(headers, [args.price_column or "", "新人价", "新人价格"], required=False)
    if price_column:
        over_price_rows = [row for row in step6_rows if to_decimal(row.get(price_column, "")) > args.price_threshold]
        step7_rows = [row for row in step6_rows if to_decimal(row.get(price_column, "")) <= args.price_threshold]
        logs.append(f"7、剔除新人价{format_decimal(args.price_threshold)}以上{len(over_price_rows)}个，剩余{len(step7_rows)}个。")
    else:
        step7_rows = step6_rows
        logs.append(f"7、未执行：源文件无“新人价”字段且未指定价格字段，继续保留上一步{len(step7_rows)}个。")

    submitted_skus = load_sku_set(args.submitted_skus, sku_column) if args.submitted_skus else set()
    step8_rows, step8_removed = remove_skus(step7_rows, sku_column, submitted_skus)
    if args.submitted_skus:
        logs.append(f"8、剔除已经提报过的sku {step8_removed}个，剩余{len(step8_rows)}个。")
    else:
        logs.append(f"8、未执行：未提供“已经提报过sku”名单，继续保留上一步{len(step8_rows)}个。")

    zhaodong_rows = [row for row in step8_rows if str(row.get(erp_column, "")).strip() == "zhaodong.103"]
    final_rows = [row for row in step8_rows if str(row.get(erp_column, "")).strip() != "zhaodong.103"]
    logs.append(f"9、剔除医药京喜的sku（zhaodong.103）{len(zhaodong_rows)}个，剩余{len(final_rows)}个。")

    csv_path = args.output_dir / f"新人价筛品结果_{args.date_tag}.csv"
    xlsx_path = args.output_dir / f"新人价筛品结果_{args.date_tag}.xlsx"
    log_path = args.output_dir / f"新人价筛品日志_{args.date_tag}.txt"
    script_copy_path = args.output_dir / f"新人价筛品自动化脚本_{args.date_tag}.py"

    write_csv(csv_path, headers, final_rows)
    xlsx_written = write_xlsx(xlsx_path, headers, final_rows, logs)
    log_path.write_text("\n".join(logs) + "\n", encoding="utf-8")
    script_copy_path.write_text(Path(__file__).read_text(encoding="utf-8"), encoding="utf-8")

    print(f"CSV={csv_path}")
    if xlsx_written:
        print(f"XLSX={xlsx_path}")
    else:
        print("XLSX_SKIPPED=openpyxl unavailable")
    print(f"LOG={log_path}")
    print(f"SCRIPT={script_copy_path}")
    print("---LOG---")
    print("\n".join(logs))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


