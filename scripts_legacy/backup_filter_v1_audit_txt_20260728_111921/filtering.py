from __future__ import annotations

import csv
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from .config import AppConfig


def detect_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "gb18030", "gbk", "cp936", "utf-16"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                file.read(8192)
            return encoding
        except UnicodeError:
            continue
    return "gb18030"


def read_table(path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows_iter = worksheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows_iter)]
            rows = []
            for row in rows_iter:
                rows.append({headers[index]: str(value or "") for index, value in enumerate(row[: len(headers)])})
            return headers, rows, "excel"
        finally:
            workbook.close()

    encoding = detect_encoding(path)
    with path.open("r", encoding=encoding, newline="") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        rows = [{key: (value or "") for key, value in row.items()} for row in reader]
    return headers, rows, encoding


def normalize_header(value: str) -> str:
    return value.replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")


def find_column(headers: list[str], candidates: list[str], required: bool = True) -> str | None:
    candidates = [candidate for candidate in candidates if candidate]
    normalized = {normalize_header(header): header for header in headers}
    for candidate in candidates:
        if normalize_header(candidate) in normalized:
            return normalized[normalize_header(candidate)]
    for header in headers:
        compact = normalize_header(header)
        if any(normalize_header(candidate) in compact for candidate in candidates):
            return header
    if required:
        raise KeyError(f"缺少必要字段：{', '.join(candidates)}")
    return None


def to_decimal(value: object) -> Decimal:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def format_decimal(value: object) -> str:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return str(int(value))
        text = format(value.normalize(), "f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    return str(value or "")


def load_sku_set(path_text: str, sku_column_name: str) -> set[str]:
    if not path_text:
        return set()
    path = Path(path_text)
    if not path.exists():
        raise FileNotFoundError(f"SKU名单文件不存在：{path}")
    headers, rows, _ = read_table(path)
    sku_column = sku_column_name if sku_column_name in headers else find_column(headers, [sku_column_name, "SKU", "sku", "SKU(含影)"], required=False)
    if not sku_column:
        sku_column = headers[0]
    return {str(row.get(sku_column, "")).strip() for row in rows if str(row.get(sku_column, "")).strip()}


def deduplicate_by_sku(rows: list[dict[str, str]], sku_column: str, metric_columns: list[str]) -> list[dict[str, str | Decimal]]:
    grouped: dict[str, dict[str, str | Decimal]] = {}
    for row in rows:
        sku = str(row.get(sku_column, "")).strip()
        if sku not in grouped:
            grouped[sku] = dict(row)
            for metric_column in metric_columns:
                grouped[sku][metric_column] = to_decimal(row.get(metric_column, ""))
        else:
            for metric_column in metric_columns:
                grouped[sku][metric_column] = to_decimal(grouped[sku].get(metric_column, "")) + to_decimal(row.get(metric_column, ""))
    return list(grouped.values())


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: format_decimal(row.get(header, "")) for header in headers})


def write_result_xlsx(path: Path, headers: list[str], rows: list[dict[str, Any]], logs: list[str]) -> None:
    workbook = Workbook(write_only=False)
    result_sheet = workbook.active
    result_sheet.title = "筛选结果"
    result_sheet.append(headers)
    for row in rows:
        result_sheet.append([format_decimal(row.get(header, "")) for header in headers])
    for cell in result_sheet[1]:
        cell.font = Font(bold=True)
    result_sheet.freeze_panes = "A2"
    for column in result_sheet.columns:
        header = str(column[0].value or "")
        result_sheet.column_dimensions[column[0].column_letter].width = min(max(len(header) + 2, 12), 28)

    log_sheet = workbook.create_sheet("筛品日志")
    for line in logs:
        log_sheet.append([line])
    log_sheet.column_dimensions["A"].width = 120
    for row in log_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    workbook.save(path)


def remove_skus(rows: list[dict[str, Any]], sku_column: str, skus: set[str]) -> tuple[list[dict[str, Any]], int]:
    if not skus:
        return rows, 0
    kept = [row for row in rows if str(row.get(sku_column, "")).strip() not in skus]
    return kept, len(rows) - len(kept)


def run_filtering(config: AppConfig) -> dict[str, Any]:
    source = Path(str(config.get("source_file")))
    if not source.exists():
        raise FileNotFoundError(f"源文件不存在：{source}")

    output_dir = config.filter_output_dir
    output_dir.mkdir(parents=True, exist_ok=True)

    headers, source_rows, source_type = read_table(source)
    sku_column = find_column(headers, [str(config.get("sku_column")), "SKU(含影)", "SKU", "sku"])
    product_name_column = find_column(headers, [str(config.get("product_name_column")), "sku名称", "商品名称"])
    mapping_column = find_column(headers, [str(config.get("mapping_column")), "1自营3POP_映射"])
    shop_column = find_column(headers, [str(config.get("shop_column")), "店铺名称"])
    department_column = find_column(headers, [str(config.get("department_column")), "1级部门名称（类目运营）", "一级部门"])
    category2_column = find_column(headers, [str(config.get("category2_column")), "商品二级分类名称"])
    erp_column = find_column(headers, [str(config.get("erp_column")), "销售员ERP帐号", "销售员ERP账号"])

    metric_columns = [find_column(headers, [name]) for name in config.get("metric_sum_columns", [])]
    keep_mapping_values = set(config.get("keep_mapping_values", []))
    keep_pop_shop_keyword = str(config.get("keep_pop_shop_keyword", ""))
    excluded_categories = set(config.get("exclude_category2", []))
    excluded_name_keywords = list(config.get("exclude_name_keywords", []))
    excluded_erp_accounts = set(config.get("exclude_erp_accounts", []))
    target_departments = list(config.get("target_departments", []))

    logs: list[str] = []
    logs.append(f"1、取数模型提取近30天动销>200，一级部门为“{'、'.join(target_departments)}”的sku获得{len(source_rows)}个sku记录（未去重）。")
    step1_rows = [row for row in source_rows if any(keyword in str(row.get(department_column, "")) for keyword in target_departments)]
    if len(step1_rows) != len(source_rows):
        logs.append(f"   源文件读取{len(source_rows)}个，一级部门过滤后剩余{len(step1_rows)}个。")

    kept_mapping_rows = [row for row in step1_rows if str(row.get(mapping_column, "")).strip() in keep_mapping_values]
    jingxi_rows = [row for row in step1_rows if str(row.get(mapping_column, "")).strip().upper() == "POP" and keep_pop_shop_keyword in str(row.get(shop_column, ""))]
    step2_rows = [row for row in step1_rows if str(row.get(mapping_column, "")).strip() in keep_mapping_values or (str(row.get(mapping_column, "")).strip().upper() == "POP" and keep_pop_shop_keyword in str(row.get(shop_column, "")))]
    logs.append(f"2、保留{'+'.join(keep_mapping_values)}（{len(kept_mapping_rows)}）、{keep_pop_shop_keyword}（{len(jingxi_rows)}），剔除{len(step1_rows) - len(step2_rows)}个，剩余{len(step2_rows)}个。")

    category_counts = Counter(str(row.get(category2_column, "")).strip() for row in step2_rows)
    step3_rows = [row for row in step2_rows if str(row.get(category2_column, "")).strip() not in excluded_categories]
    category_text = "、".join([f"“{category}”（{category_counts[category]}个）" for category in config.get("exclude_category2", [])])
    logs.append(f"3、剔除{category_text}，剩余{len(step3_rows)}个。")

    step4_rows = deduplicate_by_sku(step3_rows, sku_column, metric_columns)
    logs.append(f"4、按照sku去重，并对“{'、'.join(metric_columns)}”按sku求和，sku去重后{len(step4_rows)}个。")

    keyword_counts = {keyword: sum(1 for row in step4_rows if keyword in str(row.get(product_name_column, ""))) for keyword in excluded_name_keywords}
    step5_rows = [row for row in step4_rows if not any(keyword in str(row.get(product_name_column, "")) for keyword in excluded_name_keywords)]
    keyword_text = "、".join([f"“{keyword}”（{count}个）" for keyword, count in keyword_counts.items()])
    logs.append(f"5、删除{keyword_text}，实际剔除{len(step4_rows) - len(step5_rows)}个，剩余{len(step5_rows)}个。")

    current_success_file = str(config.get("current_success_skus_file", "") or "").strip()
    current_success_skus = load_sku_set(current_success_file, sku_column) if current_success_file else set()
    step6_rows, step6_removed = remove_skus(step5_rows, sku_column, current_success_skus)
    if current_success_file:
        logs.append(f"6、删除目前提报成功的sku {step6_removed}个，剩余{len(step6_rows)}个。")
    else:
        logs.append(f"6、未执行：未提供“目前提报成功sku”名单，继续保留上一步{len(step6_rows)}个。")

    price_column = find_column(headers, [str(config.get("price_column", "")), "新人价", "新人价格"], required=False)
    if price_column:
        threshold = to_decimal(config.get("price_threshold", 200))
        over_price_rows = [row for row in step6_rows if to_decimal(row.get(price_column, "")) > threshold]
        step7_rows = [row for row in step6_rows if to_decimal(row.get(price_column, "")) <= threshold]
        logs.append(f"7、剔除新人价{format_decimal(threshold)}以上{len(over_price_rows)}个，剩余{len(step7_rows)}个。")
    else:
        step7_rows = step6_rows
        logs.append(f"7、未执行：源文件无“新人价”字段且未指定价格字段，继续保留上一步{len(step7_rows)}个。")

    submitted_file = str(config.get("submitted_skus_file", "") or "").strip()
    submitted_skus = load_sku_set(submitted_file, sku_column) if submitted_file else set()
    step8_rows, step8_removed = remove_skus(step7_rows, sku_column, submitted_skus)
    if submitted_file:
        logs.append(f"8、剔除已经提报过的sku {step8_removed}个，剩余{len(step8_rows)}个。")
    else:
        logs.append(f"8、未执行：未提供“已经提报过sku”名单，继续保留上一步{len(step8_rows)}个。")

    erp_removed = [row for row in step8_rows if str(row.get(erp_column, "")).strip() in excluded_erp_accounts]
    final_rows = [row for row in step8_rows if str(row.get(erp_column, "")).strip() not in excluded_erp_accounts]
    logs.append(f"9、剔除指定ERP帐号（{'、'.join(excluded_erp_accounts)}）{len(erp_removed)}个，剩余{len(final_rows)}个。")

    date_tag = str(config.get("date_tag"))
    xlsx_path = output_dir / f"新人价筛品结果_{date_tag}.xlsx"
    log_path = output_dir / f"新人价筛品日志_{date_tag}.txt"
    write_result_xlsx(xlsx_path, headers, final_rows, logs)
    log_path.write_text("\n".join(logs) + "\n", encoding="utf-8")

    csv_path = None
    if bool(config.get("export_csv", False)):
        csv_path = output_dir / f"新人价筛品结果_{date_tag}.csv"
        write_csv(csv_path, headers, final_rows)

    return {"rows": len(final_rows), "xlsx": xlsx_path, "log": log_path, "csv": csv_path, "logs": logs, "source_type": source_type}
