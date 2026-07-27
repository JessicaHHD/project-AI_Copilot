from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook


def extract_skus(source: Path, source_sheet: str | None, source_column: str) -> list[str]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook[source_sheet] if source_sheet and source_sheet in workbook.sheetnames else workbook.active
        header_row = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        if source_column not in header_row:
            raise ValueError(f"未找到字段：{source_column}")
        sku_index = header_row.index(source_column)

        skus: list[str] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            value = row[sku_index]
            if value is not None and str(value).strip():
                skus.append(str(value).strip())
        return skus
    finally:
        workbook.close()


def save_sku_file(path: Path, sheet_name: str, skus: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(["sku"])
    for sku in skus:
        worksheet.append([sku])
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 22
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser(description="从筛品结果提取SKU，并按固定数量拆分为多个Excel文件")
    parser.add_argument("--source", required=True, type=Path, help="筛品结果Excel路径")
    parser.add_argument("--output-dir", required=True, type=Path, help="输出文件夹")
    parser.add_argument("--base-name", default="8月第一批查价前sku", help="输出文件名前缀")
    parser.add_argument("--source-sheet", default="筛选结果", help="源Excel工作表名称")
    parser.add_argument("--source-column", default="SKU(含影)", help="源Excel中的SKU字段名")
    parser.add_argument("--chunk-size", default=5000, type=int, help="每个拆分文件记录数")
    args = parser.parse_args()

    if args.chunk_size <= 0:
        raise ValueError("chunk-size 必须大于0")

    args.output_dir.mkdir(parents=True, exist_ok=True)
    skus = extract_skus(args.source, args.source_sheet, args.source_column)

    summary_path = args.output_dir / f"{args.base_name}.xlsx"
    save_sku_file(summary_path, "汇总", skus)

    part_paths: list[Path] = []
    for start in range(0, len(skus), args.chunk_size):
        part_number = start // args.chunk_size + 1
        part_path = args.output_dir / f"{args.base_name}_PART{part_number:02d}.xlsx"
        save_sku_file(part_path, f"PART{part_number:02d}", skus[start:start + args.chunk_size])
        part_paths.append(part_path)

    print(f"汇总文件：{summary_path}")
    print(f"SKU数量：{len(skus)}")
    print(f"拆分文件数：{len(part_paths)}")
    for path in part_paths:
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
