from __future__ import annotations

from collections import Counter
from datetime import datetime
import json
import os
from pathlib import Path
from typing import Any

from .batch import batch_date, batch_name, manifest_dir
from .config import AppConfig, PROJECT_ROOT, load_config


def text(value: Any) -> str:
    return str(value or '').strip()


def display_path(path: Path | str | None) -> str:
    if not path:
        return ''
    candidate = Path(str(path))
    try:
        return str(candidate.resolve().relative_to(PROJECT_ROOT.resolve()))
    except (OSError, ValueError):
        return str(candidate)


def latest_file(directory: Path, patterns: list[str]) -> Path | None:
    if not directory.exists():
        return None
    matches: list[Path] = []
    for pattern in patterns:
        matches.extend(path for path in directory.glob(pattern) if path.is_file() and not path.name.startswith('~$'))
    return max(matches, key=lambda path: path.stat().st_mtime) if matches else None


def read_json(path: Path | None) -> dict[str, Any]:
    if not path:
        return {}
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception:
        return {}


def read_recent_lines(path: Path, limit: int = 80) -> list[str]:
    for encoding in ('utf-8-sig', 'utf-8', 'gbk'):
        try:
            lines = path.read_text(encoding=encoding, errors='ignore').splitlines()
            return [line.strip() for line in lines if line.strip()][-limit:]
        except Exception:
            continue
    return []


def read_workbook_rows(path: Path, max_rows: int = 50000) -> tuple[list[str], list[list[str]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError('当前 Python 环境未安装 openpyxl，暂不能读取 Excel 汇总指标。') from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        iterator = worksheet.iter_rows(values_only=True)
        headers = [text(cell) for cell in next(iterator, [])]
        rows: list[list[str]] = []
        for row_index, row in enumerate(iterator):
            if row_index >= max_rows:
                break
            values = [text(cell) for cell in row]
            if any(values):
                rows.append(values)
        return headers, rows
    finally:
        workbook.close()


def latest_manifest_file(config: AppConfig) -> Path | None:
    directory = manifest_dir(config)
    if not directory.exists():
        return None
    candidates = [path for path in directory.glob(f'{batch_name(config)}_*.json') if path.is_file()]
    if not candidates:
        candidates = [path for path in directory.glob('*.json') if path.is_file()]
    return max(candidates, key=lambda path: path.stat().st_mtime) if candidates else None


def scan_latest_outputs(config: AppConfig | None = None) -> dict[str, Any]:
    config = config or load_config()
    output_root = config.output_root
    final_dir = output_root / '最终结果'
    price_dir = config.download_output_dir
    split_dir = config.split_output_dir
    submit_dir = Path(str(config.get('registration_submit_dir') or output_root / '提报文件'))
    status_export_dir = Path(str(config.get('registration_status_export_dir') or output_root / '提报情况导出'))
    suspicious_dir = price_dir / '疑似误下载'
    if not suspicious_dir.exists() and price_dir.exists():
        candidates = [path for path in price_dir.rglob('*疑似误下载*') if path.is_dir()]
        if candidates:
            suspicious_dir = max(candidates, key=lambda path: path.stat().st_mtime)
    debug_dir = PROJECT_ROOT / 'logs' / 'debug_backend'
    if not debug_dir.exists() and config.log_root.exists():
        candidates = [path for path in config.log_root.glob('debug*') if path.is_dir()]
        debug_dir = max(candidates, key=lambda path: path.stat().st_mtime) if candidates else config.log_root
    return {
        'output_root': output_root, 'final_dir': final_dir, 'price_dir': price_dir, 'split_dir': split_dir,
        'submit_dir': submit_dir, 'status_export_dir': status_export_dir,
        'pricing_file': latest_file(final_dir, ['*筛品查价表*.xlsx', '*新人价最终结果*.xlsx']),
        'failed_sku_file': latest_file(final_dir, ['*查价失败*SKU*.xlsx', '*查价失败*.xlsx', '*未匹配*.xlsx']),
        'status_merge_file': latest_file(final_dir, ['*提报情况整理表*.xlsx', '*提报整理*.xlsx']),
        'status_export_file': latest_file(status_export_dir, ['*.xlsx', '*.csv']),
        'manifest_file': latest_manifest_file(config), 'suspicious_dir': suspicious_dir, 'debug_dir': debug_dir,
        'has_price_exports': latest_file(price_dir, ['*.xlsx', '*.csv']) is not None,
        'has_split_files': latest_file(split_dir, ['*.xlsx', '*.csv']) is not None,
        'has_submit_parts': latest_file(submit_dir, ['*PART*.xlsx']) is not None,
    }


def load_batch_status(config: AppConfig | None = None, outputs: dict[str, Any] | None = None) -> list[dict[str, Any]]:
    config = config or load_config()
    outputs = outputs or scan_latest_outputs(config)
    manifest = read_json(outputs.get('manifest_file'))
    prepare_done = bool(outputs.get('manifest_file')) or bool(manifest.get('steps')) or outputs['has_split_files']
    pricing_done = bool(outputs.get('pricing_file'))
    pricing_ready = pricing_done or outputs['has_price_exports'] or outputs['has_split_files']
    parts_done = outputs['has_submit_parts']
    export_done = bool(outputs.get('status_export_file'))
    merge_done = bool(outputs.get('status_merge_file'))
    return [
        {'id': 'prepare', 'label': '准备任务', 'status': 'completed' if prepare_done else 'need_user', 'description': '已识别配置、批次目录和批次清单。' if prepare_done else '请先确认配置文件、源表和批次清单是否准备好。'},
        {'id': 'pricing', 'label': '筛品查价', 'status': 'completed' if pricing_done else ('need_user' if pricing_ready else 'not_started'), 'description': '已找到新人价整合结果表。' if pricing_done else '已找到部分筛品或查价产物，等待生成最终整合表。'},
        {'id': 'confirm', 'label': '业务确认', 'status': 'completed' if parts_done else ('need_user' if pricing_ready else 'not_started'), 'description': '提报 PART 文件已生成。' if parts_done else '请业务确认参与报名、不报名和需复核的 SKU。'},
        {'id': 'submit', 'label': '后台提报', 'status': 'completed' if export_done or merge_done else ('need_user' if parts_done else 'not_started'), 'description': '已找到后台导出的提报情况文件。' if export_done else '当前看板不执行真实提报；请在命令行工具中按需操作。'},
        {'id': 'merge', 'label': '审核回填', 'status': 'completed' if merge_done else ('need_user' if export_done else 'not_started'), 'description': '已生成提报情况整理表。' if merge_done else '等待审核结果导出后生成提报情况整理表。'},
    ]


def count_rows_containing(rows: list[list[str]], keywords: list[str]) -> int:
    keywords = [keyword for keyword in keywords if keyword]
    return sum(1 for row in rows if any(keyword in ' '.join(row) for keyword in keywords))


def pool_overview(status_file: Path | None, config: AppConfig) -> dict[str, Any]:
    empty = {'available': False, 'message': '暂未找到提报情况整理表，品池概览先显示占位。', 'source_file': '', 'last_submitted_count': None, 'success_count': None, 'failed_count': None, 'resubmittable_count': None, 'exit_count': None, 'failure_reasons': []}
    if not status_file or not status_file.exists():
        return empty
    try:
        headers, rows = read_workbook_rows(status_file)
    except Exception as exc:
        return {**empty, 'message': str(exc), 'source_file': str(status_file)}
    reasons = failure_reasons(headers, rows, config)
    return {
        'available': True, 'message': '已读取最新提报情况整理表。', 'source_file': str(status_file),
        'last_submitted_count': len(rows),
        'success_count': count_rows_containing(rows, ['成功', '通过', '已通过', '报名成功', '已报名']),
        'failed_count': count_rows_containing(rows, ['失败', '拒绝', '不通过', '未通过', '转化不满足', '删促']),
        'resubmittable_count': count_rows_containing(rows, [str(item) for item in config.get('resubmit_success_keywords', [])]),
        'exit_count': count_rows_containing(rows, [str(item) for item in config.get('resubmit_exit_keywords', [])]),
        'failure_reasons': reasons,
    }


def failure_reasons(headers: list[str], rows: list[list[str]], config: AppConfig) -> list[dict[str, Any]]:
    columns = [index for index, header in enumerate(headers) if any(key in header for key in ('原因', '失败', '拒绝', '状态', '结果', '备注'))]
    keywords = [*[str(item) for item in config.get('resubmit_success_keywords', [])], '活动规则不满足', '价格不满足', '库存不足', '审核拒绝', '未通过', '失败']
    counter: Counter[str] = Counter()
    for row in rows:
        row_text = ' '.join(row)
        keyword = next((item for item in keywords if item and item in row_text), '')
        if keyword:
            counter[keyword] += 1
            continue
        for column in columns:
            reason = row[column] if column < len(row) else ''
            if reason and reason not in {'成功', '通过', '已通过', '报名成功', '已完成'}:
                counter[reason[:40]] += 1
                break
    return [{'reason': reason, 'count': count} for reason, count in counter.most_common(3)]


def extract_erp_text(pricing_file: Path | None, erp_column: str) -> str:
    if not pricing_file or not pricing_file.exists():
        return ''
    try:
        headers, rows = read_workbook_rows(pricing_file, max_rows=30000)
    except Exception:
        return ''
    erp_index = next((index for index, header in enumerate(headers) if header == erp_column or 'ERP' in header.upper()), -1)
    if erp_index < 0:
        return ''
    return ';'.join(sorted({row[erp_index] for row in rows if erp_index < len(row) and row[erp_index]}))


def latest_log_status(config: AppConfig, outputs: dict[str, Any]) -> dict[str, Any]:
    log_files: list[Path] = []
    if config.log_root.exists():
        log_files.extend(path for path in config.log_root.glob('*.txt') if path.is_file())
        log_files.extend(path for path in config.log_root.glob('*.log') if path.is_file())
    latest_log = max(log_files, key=lambda path: path.stat().st_mtime) if log_files else None
    message = '暂未找到日志文件。'
    exception = '暂无异常'
    if latest_log:
        lines = read_recent_lines(latest_log)
        message = lines[-1] if lines else message
        errors = [line for line in lines if any(key in line.lower() for key in ('error', 'exception', 'traceback', '失败', '异常', '报错'))]
        exception = errors[-1] if errors else exception
    return {
        'latest_log': {'path': str(latest_log or ''), 'display_path': display_path(latest_log), 'message': message, 'updated_at': datetime.fromtimestamp(latest_log.stat().st_mtime).strftime('%Y-%m-%d %H:%M:%S') if latest_log else ''},
        'current_exception': exception, 'debug_dir': str(outputs['debug_dir']), 'log_dir': str(config.log_root), 'suspicious_dir': str(outputs['suspicious_dir']),
    }


def next_action(stages: list[dict[str, Any]]) -> tuple[str, str]:
    suggestions = {'prepare': '请检查配置、源表和批次清单。', 'pricing': '请通过现有命令行工具完成筛品查价整合。', 'confirm': '请复制 ERP 文本并发送业务确认，收到确认表后再继续。', 'submit': '如已确认业务表，请通过命令行工具执行后台提报。', 'merge': '请导出审核结果并通过命令行工具生成整理表。'}
    for stage in stages:
        if stage['status'] in {'need_user', 'error'}:
            return stage['label'], suggestions.get(stage['id'], stage['description'])
    return '批次完成', '核心产物已找到，可复盘异常或准备下一批次。'


def load_gui_context() -> dict[str, Any]:
    config = load_config()
    outputs = scan_latest_outputs(config)
    stages = load_batch_status(config, outputs)
    current_stage, next_step = next_action(stages)
    manifest = read_json(outputs.get('manifest_file'))
    config_path = os.environ.get('NEWCOMER_TOOL_CONFIG', 'config.yaml')
    environment = '测试环境' if 'test' in Path(config_path).name.lower() else '本地环境'
    return {'batch_name': manifest.get('batch_name') or batch_name(config), 'environment': environment, 'batch_date': text(config.get('batch_date')) or manifest.get('batch_date') or batch_date(config), 'output_dir': str(config.output_root), 'log_dir': str(config.log_root), 'current_stage': current_stage, 'next_step': next_step, 'manifest_status': '已找到' if outputs.get('manifest_file') else '暂未找到', 'manifest_path': str(outputs.get('manifest_file') or '')}


def file_status(path: Path | None, directory_allowed: bool = False) -> str:
    if path and path.exists():
        return 'empty' if path.is_dir() and not any(path.iterdir()) else 'found'
    return 'empty' if directory_allowed else 'missing'


def load_dashboard() -> dict[str, Any]:
    config = load_config()
    outputs = scan_latest_outputs(config)
    erp_text = extract_erp_text(outputs['pricing_file'], str(config.get('erp_column') or 'ERP'))
    erp_count = len([item for item in erp_text.split(';') if item])
    key_files = [
        {'id': 'pricing_file', 'label': '新人价整合表', 'status': file_status(outputs['pricing_file']), 'path': str(outputs['pricing_file'] or ''), 'display_path': display_path(outputs['pricing_file']), 'action': 'open'},
        {'id': 'failed_sku', 'label': '查价失败 SKU', 'status': file_status(outputs['failed_sku_file']), 'path': str(outputs['failed_sku_file'] or ''), 'display_path': display_path(outputs['failed_sku_file']), 'action': 'open'},
        {'id': 'erp_text', 'label': 'ERP 文本', 'status': 'available' if erp_text else 'missing', 'path': erp_text, 'display_path': f'已识别 {erp_count} 个 ERP' if erp_text else '暂未从整合表识别 ERP', 'action': 'copy'},
        {'id': 'part_dir', 'label': '提报 PART 文件目录', 'status': 'found' if outputs['submit_dir'].exists() else 'missing', 'path': str(outputs['submit_dir']), 'display_path': display_path(outputs['submit_dir']), 'action': 'open'},
        {'id': 'status_merge', 'label': '提报情况整理表', 'status': file_status(outputs['status_merge_file']), 'path': str(outputs['status_merge_file'] or ''), 'display_path': display_path(outputs['status_merge_file']), 'action': 'open'},
        {'id': 'manifest', 'label': '批次清单', 'status': file_status(outputs['manifest_file']), 'path': str(outputs['manifest_file'] or ''), 'display_path': display_path(outputs['manifest_file']), 'action': 'open'},
        {'id': 'log_dir', 'label': '日志目录', 'status': 'found' if config.log_root.exists() else 'missing', 'path': str(config.log_root), 'display_path': display_path(config.log_root), 'action': 'open'},
        {'id': 'suspicious', 'label': '疑似误下载目录', 'status': file_status(outputs['suspicious_dir'], directory_allowed=True), 'path': str(outputs['suspicious_dir']), 'display_path': display_path(outputs['suspicious_dir']), 'action': 'open'},
    ]
    return {'context': load_gui_context(), 'pool_overview': pool_overview(outputs['status_merge_file'], config), 'stages': load_batch_status(config, outputs), 'key_files': key_files, 'log_status': latest_log_status(config, outputs)}
