from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from .batch import registration_status_output_path
from .config import AppConfig
from .final_pricing import find_column, normalize

BUSINESS_PARTICIPATION_COLUMN = "业务是否参与报名"
ADDED_COLUMNS = ["提报记录来源", BUSINESS_PARTICIPATION_COLUMN, "提报状态_最终", "提报结果标签", "提报详情", "提报记录数"]
STATUS_ADDED_COLUMNS = ["提报记录来源", "提报状态_最终", "提报结果标签", "提报详情", "提报记录数"]
LEGACY_CONFIRM_COLUMNS = ["是否退出_采销确认"]
SUMMARY_SHEETS = ["提报情况原始", "提报情况归并", "非本轮主表提报记录", "业务确认明细", "提报汇总", "提报整理日志"]
SUCCESS_STATUS = "提报成功"
FAILED_STATUS = "提报失败"
MISSING_STATUS = "未提报/未匹配"
MAIN_SOURCE = "本轮主表SKU"
EXTRA_SOURCE = "后台额外提报SKU"
JOIN_STATUS_YES = "参与报名"
JOIN_STATUS_NO = "不报名"
JOIN_STATUS_UNKNOWN = "未确认"
NO_VALUES = {"否", "n", "no", "不", "不参加", "不报名", "不提报", "退出", "取消", "放弃", "false", "0"}
EXIT_VALUES = {"是", "y", "yes", "true", "1", "退出", "不参加", "不报名", "不提报", "取消", "放弃"}


@dataclass
class StatusRecord:
    sku: str
    status: str
    detail: str
    promo_status: str
    label: str
    effective_status: str
    row_order: int
    time_value: datetime | None
    raw: dict[str, Any]


@dataclass
class MergedStatus:
    sku: str
    final_status: str
    result_label: str
    detail: str
    promo_status: str
    record_count: int
    source_row_order: int
    source_time: datetime | None


def raw_value_by_candidates(row: dict[str, Any], candidates: list[str]) -> Any:
    if not row:
        return ""
    normalized = {compact(key): key for key in row.keys()}
    for candidate in candidates:
        key = normalized.get(compact(candidate))
        if key is not None and text(row.get(key)):
            return row.get(key)
    for key, value in row.items():
        key_text = compact(key)
        if any(compact(candidate) in key_text for candidate in candidates) and text(value):
            return value
    return ""


def build_status_record_map(records: list[StatusRecord], merged_by_sku: dict[str, MergedStatus]) -> dict[str, StatusRecord]:
    chosen_by_sku: dict[str, StatusRecord] = {}
    for record in records:
        merged = merged_by_sku.get(record.sku)
        if not merged or record.row_order != merged.source_row_order:
            continue
        chosen_by_sku[record.sku] = record
    return chosen_by_sku


def fill_extra_row_from_status(sheet, row_index: int, headers: list[str], sku_col_name: str, sku: str, raw: dict[str, Any]) -> str:
    candidate_map = {
        sku_col_name: ["skuid", "skuId", "SKU", "sku", "SKU(含影)"],
        "sku名称": ["sku名称", "商品名称", "商品标题", "商品信息", "商品简称"],
        "商品名称": ["商品名称", "sku名称", "商品标题", "商品信息", "商品简称"],
        "销售员ERP帐号": ["销售员ERP帐号", "销售员ERP账号", "采销ERP", "运营ERP", "提报人", "创建人", "销售员ERP"],
        "销售员ERP账号": ["销售员ERP账号", "销售员ERP帐号", "采销ERP", "运营ERP", "提报人", "创建人", "销售员ERP"],
        "店铺名称": ["店铺名称", "店铺"],
        "1级部门名称（类目运营）": ["1级部门名称（类目运营）", "一级部门", "部门", "事业部"],
        "商品二级分类名称": ["商品二级分类名称", "二级分类", "类目", "商品类目"],
    }
    filled = 0
    for column_index, header in enumerate(headers, start=1):
        if header in ADDED_COLUMNS:
            continue
        if header == sku_col_name:
            sheet.cell(row_index, column_index, sku)
            filled += 1
            continue
        value = raw.get(header, "")
        if not text(value):
            value = raw_value_by_candidates(raw, candidate_map.get(header, [header]))
        if text(value):
            sheet.cell(row_index, column_index, value)
            filled += 1
    return "后台导出表" if filled > 1 else "仅识别SKU"


def normalize_confirm_value(value: Any) -> str:
    return text(value).replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "").lower()


def business_participation(field_name: str, raw_value: Any, found: bool = True) -> tuple[str, str]:
    if not found:
        return JOIN_STATUS_UNKNOWN, "业务确认表未找到该SKU"
    value = normalize_confirm_value(raw_value)
    if not value:
        return JOIN_STATUS_YES, "空白默认参与报名"
    field = compact(field_name)
    if "退出" in field:
        if value in EXIT_VALUES or any(keyword in value for keyword in ["不参加", "不报名", "不提报", "退出", "取消", "放弃"]):
            return JOIN_STATUS_NO, "业务确认字段表示退出/不报名"
        return JOIN_STATUS_YES, "未命中退出/不报名关键词"
    if value in NO_VALUES or any(keyword in value for keyword in ["不参加", "不报名", "不提报", "退出", "取消", "放弃"]):
        return JOIN_STATUS_NO, "业务确认字段表示不报名"
    return JOIN_STATUS_YES, "未命中不报名关键词"


def text(value: Any) -> str:
    return str(value or "").strip()


def sku_text(value: Any) -> str:
    raw = text(value)
    if raw.endswith(".0"):
        raw = raw[:-2]
    return raw


def compact(value: Any) -> str:
    return normalize(text(value))


def latest_file(directory: Path, pattern: str, description: str, fallback_patterns: list[str] | None = None) -> Path:
    if not directory.exists():
        raise FileNotFoundError(f"{description}目录不存在：{directory}")
    patterns = [pattern, *(fallback_patterns or [])]
    for current_pattern in patterns:
        candidates = sorted(
            (path for path in directory.glob(current_pattern) if path.is_file() and not path.name.startswith("~$")),
            key=lambda item: item.stat().st_mtime,
            reverse=True,
        )
        if candidates:
            return candidates[0]
    raise FileNotFoundError(f"未找到{description}：{directory / pattern}")


def resolve_input_files(config: AppConfig) -> tuple[Path, Path, Path]:
    status_config = text(config.get("registration_status_file"))
    final_config = text(config.get("registration_status_final_file"))
    confirm_config = text(config.get("registration_status_confirm_file"))

    status_file = Path(status_config) if status_config else latest_file(
        Path(str(config.get("registration_status_export_dir") or config.output_root / "提报情况导出")),
        "*.xlsx",
        "提报情况导出文件",
    )
    final_file = Path(final_config) if final_config else latest_file(
        config.output_root / "最终结果",
        "*筛品查价表*.xlsx",
        "最终结果文件",
        fallback_patterns=["新人价最终结果_*.xlsx"],
    )
    confirm_file = Path(confirm_config) if confirm_config else latest_file(
        Path(str(config.get("registration_input_dir") or config.output_root.parent / "input" / "提报确认表")),
        "*业务已确认表_*.xlsx",
        "业务确认表",
        fallback_patterns=["*新人价采销确认表*.xlsx", "*.xlsx"],
    )
    for label, path in [("提报情况导出文件", status_file), ("最终结果文件", final_file), ("采销确认表", confirm_file)]:
        if not path.exists():
            raise FileNotFoundError(f"{label}不存在：{path}")
    return status_file, final_file, confirm_file


def worksheet_rows(path: Path) -> tuple[str, list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for worksheet in workbook.worksheets:
            if hasattr(worksheet, "reset_dimensions"):
                worksheet.reset_dimensions()
            rows_iter = worksheet.iter_rows(max_col=200, values_only=True)
            headers: list[str] = []
            for row_number, row in enumerate(rows_iter, start=1):
                current = [text(value) for value in row]
                if any(current):
                    headers = current
                    break
                if row_number >= 50:
                    break
            if not headers:
                continue
            records: list[dict[str, Any]] = []
            empty_streak = 0
            for data_row in rows_iter:
                record = {headers[index]: data_row[index] if index < len(data_row) else None for index in range(len(headers))}
                if any(text(value) for value in record.values()):
                    records.append(record)
                    empty_streak = 0
                else:
                    empty_streak += 1
                    if empty_streak >= 200:
                        break
            return worksheet.title, headers, records
    finally:
        workbook.close()
    return "", [], []


def parse_time(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    raw = text(value)
    if not raw:
        return None
    for fmt in ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d"]:
        try:
            return datetime.strptime(raw, fmt)
        except ValueError:
            continue
    return None


def classify_failure(detail: str) -> str:
    value = compact(detail)
    if "交叉时间内已存在" in value:
        return "叠加促销过多"
    if "严重错误" in value:
        return "重复提报"
    if "不满足促销价低于100天最低价" in value:
        return "不满足促销价低于100天最低价"
    if "近30天日均拉新量" in value:
        return "日均拉新不满足"
    if "不符合30天转化用户数" in value:
        return "30天转化不满足"
    return text(detail)


def is_success_status(status: str) -> bool:
    value = compact(status)
    return "成功" in value and "失败" not in value


def status_time_column(headers: list[str]) -> str | None:
    candidates = ["提报时间", "报名时间", "更新时间", "创建时间", "提交时间", "导入时间", "操作时间"]
    try:
        return find_column(headers, candidates, required=False)
    except KeyError:
        return None


def read_status_records(path: Path) -> tuple[list[str], list[dict[str, Any]], list[StatusRecord], list[str]]:
    sheet_name, headers, rows = worksheet_rows(path)
    if not headers or not rows:
        raise ValueError(f"提报情况导出文件无有效数据：{path}")
    sku_col = find_column(headers, ["skuid", "skuId", "SKU", "sku", "SKU(含影)"])
    status_col = find_column(headers, ["提报状态", "报名状态"])
    detail_col = find_column(headers, ["详情", "失败原因", "原因"])
    promo_col = find_column(headers, ["促销状态", "补贴状态"])
    time_col = status_time_column(headers)

    records: list[StatusRecord] = []
    for index, row in enumerate(rows, start=2):
        sku = sku_text(row.get(sku_col))
        if not sku:
            continue
        status = text(row.get(status_col))
        detail = text(row.get(detail_col))
        promo_status = text(row.get(promo_col))
        label = promo_status or "提报成功"
        effective_status = SUCCESS_STATUS if is_success_status(status) else FAILED_STATUS
        if effective_status == FAILED_STATUS:
            label = classify_failure(detail)
            if label == "重复提报":
                effective_status = SUCCESS_STATUS
        records.append(StatusRecord(
            sku=sku,
            status=status,
            detail=detail,
            promo_status=promo_status,
            label=label,
            effective_status=effective_status,
            row_order=index,
            time_value=parse_time(row.get(time_col)) if time_col else None,
            raw=row,
        ))
    if not records:
        raise ValueError(f"提报情况导出文件未读取到有效 SKU：{path}")
    logs = [
        f"提报情况文件：{path}",
        f"提报情况sheet：{sheet_name}",
        f"识别字段：SKU={sku_col}，提报状态={status_col}，详情={detail_col}，促销状态={promo_col}，时间字段={time_col or '未识别，按行顺序'}",
        f"提报情况原始有效行数：{len(records)}",
    ]
    return headers, rows, records, logs


def newer_record(record: StatusRecord, current: StatusRecord | None) -> bool:
    if current is None:
        return True
    if record.time_value and current.time_value:
        return record.time_value > current.time_value
    if record.time_value and not current.time_value:
        return True
    if not record.time_value and current.time_value:
        return False
    return record.row_order > current.row_order


def merge_status_records(records: list[StatusRecord]) -> tuple[dict[str, MergedStatus], list[dict[str, Any]], list[str]]:
    grouped: dict[str, list[StatusRecord]] = defaultdict(list)
    for record in records:
        grouped[record.sku].append(record)

    merged: dict[str, MergedStatus] = {}
    rows: list[dict[str, Any]] = []
    success_count = 0
    fail_count = 0
    duplicate_sku_count = 0

    for sku, items in grouped.items():
        if len(items) > 1:
            duplicate_sku_count += 1
        success_items = [item for item in items if item.effective_status == SUCCESS_STATUS]
        if success_items:
            chosen: StatusRecord | None = None
            for item in success_items:
                if newer_record(item, chosen):
                    chosen = item
            assert chosen is not None
            final_status = SUCCESS_STATUS
            success_count += 1
        else:
            chosen = None
            for item in items:
                if newer_record(item, chosen):
                    chosen = item
            assert chosen is not None
            final_status = FAILED_STATUS
            fail_count += 1
        merged_item = MergedStatus(
            sku=sku,
            final_status=final_status,
            result_label=chosen.label,
            detail=chosen.detail,
            promo_status=chosen.promo_status,
            record_count=len(items),
            source_row_order=chosen.row_order,
            source_time=chosen.time_value,
        )
        merged[sku] = merged_item
        rows.append({
            "skuid": sku,
            "提报状态_最终": merged_item.final_status,
            "提报结果标签": merged_item.result_label,
            "提报详情": merged_item.detail,
            "促销状态_原始": merged_item.promo_status,
            "提报记录数": merged_item.record_count,
            "采用记录行号": merged_item.source_row_order,
            "采用记录时间": merged_item.source_time.strftime("%Y-%m-%d %H:%M:%S") if merged_item.source_time else "",
        })
    rows.sort(key=lambda row: row["skuid"])
    logs = [
        f"提报情况去重SKU数：{len(merged)}",
        f"重复SKU数：{duplicate_sku_count}",
        f"归并后成功SKU数：{success_count}",
        f"归并后失败SKU数：{fail_count}",
    ]
    return merged, rows, logs


def read_business_confirmation_map(path: Path) -> tuple[dict[str, dict[str, str]], list[dict[str, str]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        if hasattr(worksheet, "reset_dimensions"):
            worksheet.reset_dimensions()
        all_rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    for row_index, row in enumerate(all_rows):
        headers = [text(value) for value in row]
        if not any(headers):
            continue
        try:
            sku_index = headers.index(find_column(headers, ["SKU", "skuid", "skuId", "SKU(含影)"]))
            exit_header = find_column(headers, ["是否退出", "是否报名", "是否登记", "是否参加", "是否提报"], required=False)
        except Exception:
            continue
        if exit_header is None:
            continue
        exit_index = headers.index(exit_header)
        result: dict[str, dict[str, str]] = {}
        detail_rows: list[dict[str, str]] = []
        for data_row in all_rows[row_index + 1:]:
            sku = sku_text(data_row[sku_index] if sku_index < len(data_row) else "")
            if not sku:
                continue
            raw_value = text(data_row[exit_index] if exit_index < len(data_row) else "")
            status, note = business_participation(exit_header, raw_value)
            item = {
                "SKU": sku,
                "业务确认字段": exit_header,
                "业务确认原值": raw_value,
                BUSINESS_PARTICIPATION_COLUMN: status,
                "说明": note,
            }
            result[sku] = item
            detail_rows.append(item)
        logs = [
            f"采销确认表：{path}",
            f"识别表头行：第 {row_index + 1} 行",
            f"识别字段：SKU={headers[sku_index]}，是否退出/报名={exit_header}",
            f"采销确认SKU数：{len(result)}",
        ]
        return result, detail_rows, logs
    raise ValueError(f"采销确认表未识别到 SKU 和 是否退出/是否报名 字段：{path}")


def confirmation_status(confirm_map: dict[str, dict[str, str]], sku: str) -> str:
    item = confirm_map.get(sku)
    if not item:
        return JOIN_STATUS_UNKNOWN
    return item.get(BUSINESS_PARTICIPATION_COLUMN, JOIN_STATUS_UNKNOWN)


def remove_existing_columns(worksheet, headers: list[str], columns: list[str]) -> list[str]:
    for header in columns:
        while header in headers:
            index = headers.index(header) + 1
            worksheet.delete_cols(index)
            headers.pop(index - 1)
    return headers


def delete_sheet_if_exists(workbook, title: str) -> None:
    if title in workbook.sheetnames:
        del workbook[title]


def append_rows_sheet(workbook, title: str, headers: list[str], rows: list[dict[str, Any]]) -> None:
    delete_sheet_if_exists(workbook, title)
    sheet = workbook.create_sheet(title)
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header, "") for header in headers])
    style_sheet(sheet)


def read_existing_rows_sheet(workbook, title: str) -> list[dict[str, Any]]:
    if title not in workbook.sheetnames:
        return []
    sheet = workbook[title]
    headers = [text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    if not any(headers):
        return []
    rows: list[dict[str, Any]] = []
    for row_index in range(2, sheet.max_row + 1):
        row = {headers[column - 1]: sheet.cell(row_index, column).value for column in range(1, len(headers) + 1) if headers[column - 1]}
        if any(text(value) for value in row.values()):
            rows.append(row)
    return rows


def style_sheet(sheet) -> None:
    # 大表只设置表头和列宽，避免遍历数万行导致写入很慢。
    if sheet.max_row >= 1:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            sheet.column_dimensions[cell.column_letter].width = min(max(len(text(cell.value)) + 2, 12), 40)
    sheet.freeze_panes = "A2"


def build_summary_rows(final_rows: list[dict[str, Any]], merged_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    status_counter = Counter(row.get("提报状态_最终", "") for row in final_rows)
    label_by_status: dict[str, Counter[str]] = defaultdict(Counter)
    for row in final_rows:
        status = text(row.get("提报状态_最终"))
        label = text(row.get("提报结果标签"))
        if status and label:
            label_by_status[status][label] += 1

    rows: list[dict[str, Any]] = []
    for status in [SUCCESS_STATUS, FAILED_STATUS, MISSING_STATUS]:
        rows.append({"行标签": status, "计数项skuid": status_counter.get(status, 0), "层级": 0})
        for label, count in label_by_status[status].most_common():
            rows.append({"行标签": label, "计数项skuid": count, "层级": 1})
    rows.append({"行标签": "后台提报情况去重SKU", "计数项skuid": len(merged_rows), "层级": 0})
    return rows


def append_summary_sheet(workbook, rows: list[dict[str, Any]]) -> None:
    delete_sheet_if_exists(workbook, "提报汇总")
    sheet = workbook.create_sheet("提报汇总")
    sheet.append(["行标签", "计数项skuid"])
    for row in rows:
        sheet.append([row.get("行标签", ""), row.get("计数项skuid", "")])
        excel_row = sheet.max_row
        level = int(row.get("层级") or 0)
        if level:
            sheet.row_dimensions[excel_row].outlineLevel = level
            sheet.cell(excel_row, 1).alignment = Alignment(indent=level, vertical="top")
        else:
            sheet.cell(excel_row, 1).font = Font(bold=True)
            sheet.cell(excel_row, 2).font = Font(bold=True)
    sheet.sheet_properties.outlinePr.summaryBelow = False
    style_sheet(sheet)
    sheet.column_dimensions["A"].width = 100
    sheet.column_dimensions["B"].width = 16


def safe_filename_part(value: str, max_length: int = 80) -> str:
    invalid_chars = '<>:"/\\|?*'
    cleaned = "".join("_" if char in invalid_chars else char for char in text(value)).strip(" ._")
    return (cleaned or "最近导出")[:max_length]


def unique_path(path: Path) -> Path:
    if not path.exists():
        return path
    for index in range(2, 1000):
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
    raise FileExistsError(f"输出文件重名过多，请清理目录后重试：{path.parent}")


def output_path_for(config: AppConfig, status_file: Path, output_dir: Path) -> Path:
    return registration_status_output_path(config, output_dir, status_file)


def run_registration_status_merge(config: AppConfig) -> dict[str, Any]:
    status_file, final_file, confirm_file = resolve_input_files(config)
    status_headers, status_rows, status_records, status_logs = read_status_records(status_file)
    merged_by_sku, merged_rows, merge_logs = merge_status_records(status_records)
    status_record_by_sku = build_status_record_map(status_records, merged_by_sku)
    confirm_map, confirmation_detail_rows, confirm_logs = read_business_confirmation_map(confirm_file)

    workbook = load_workbook(final_file)
    if "筛选结果" not in workbook.sheetnames:
        workbook.close()
        raise ValueError(f"最终结果文件缺少“筛选结果”sheet：{final_file}")
    sheet = workbook["筛选结果"]
    headers = [text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    sku_col_name = find_column(headers, [str(config.get("sku_column")), "SKU(含影)", "skuid", "skuId", "SKU", "sku"])
    existing_detail_rows = read_existing_rows_sheet(workbook, "业务确认明细")
    headers = remove_existing_columns(sheet, headers, [*STATUS_ADDED_COLUMNS, *LEGACY_CONFIRM_COLUMNS])
    sku_col_index = headers.index(sku_col_name) + 1
    business_col_name = find_column(headers, [BUSINESS_PARTICIPATION_COLUMN], required=False)
    business_col_index = headers.index(business_col_name) + 1 if business_col_name else None

    start_col = len(headers) + 1
    output_columns = list(STATUS_ADDED_COLUMNS)
    if business_col_index is None:
        output_columns.insert(1, BUSINESS_PARTICIPATION_COLUMN)
        business_col_index = start_col + output_columns.index(BUSINESS_PARTICIPATION_COLUMN)
    for offset, header in enumerate(output_columns):
        cell = sheet.cell(1, start_col + offset, header)
        cell.font = Font(bold=True)
    final_rows_for_summary: list[dict[str, Any]] = []
    main_skus: set[str] = set()
    matched = 0
    missing = 0
    for row_index in range(2, sheet.max_row + 1):
        sku = sku_text(sheet.cell(row_index, sku_col_index).value)
        if sku:
            main_skus.add(sku)
        merged = merged_by_sku.get(sku)
        existing_participation = text(sheet.cell(row_index, business_col_index).value) if business_col_index else ""
        participation = existing_participation or confirmation_status(confirm_map, sku)
        if business_col_index:
            sheet.cell(row_index, business_col_index, participation)
        if merged:
            matched += 1
            row_values = {
                "提报记录来源": MAIN_SOURCE,
                BUSINESS_PARTICIPATION_COLUMN: participation,
                "提报状态_最终": merged.final_status,
                "提报结果标签": merged.result_label,
                "提报详情": merged.detail,
                "提报记录数": merged.record_count,
            }
        else:
            missing += 1
            row_values = {
                "提报记录来源": MAIN_SOURCE,
                BUSINESS_PARTICIPATION_COLUMN: participation,
                "提报状态_最终": MISSING_STATUS,
                "提报结果标签": "",
                "提报详情": "",
                "提报记录数": 0,
            }
        for offset, header in enumerate(output_columns):
            sheet.cell(row_index, start_col + offset, row_values.get(header, ""))
        final_rows_for_summary.append({
            "skuid": sku,
            **row_values,
        })

    extra_status_rows: list[dict[str, Any]] = []
    extra_skus = sorted(sku for sku in merged_by_sku if sku and sku not in main_skus)
    for sku in extra_skus:
        merged = merged_by_sku[sku]
        raw_record = status_record_by_sku.get(sku)
        raw = raw_record.raw if raw_record else {}
        row_index = sheet.max_row + 1
        fill_source = fill_extra_row_from_status(sheet, row_index, headers, sku_col_name, sku, raw)
        row_values = {
            "提报记录来源": EXTRA_SOURCE,
            BUSINESS_PARTICIPATION_COLUMN: JOIN_STATUS_UNKNOWN,
            "提报状态_最终": merged.final_status,
            "提报结果标签": merged.result_label,
            "提报详情": merged.detail,
            "提报记录数": merged.record_count,
        }
        if business_col_index:
            sheet.cell(row_index, business_col_index, row_values[BUSINESS_PARTICIPATION_COLUMN])
        for offset, header in enumerate(output_columns):
            sheet.cell(row_index, start_col + offset, row_values.get(header, ""))
        final_rows_for_summary.append({
            "skuid": sku,
            **row_values,
        })
        extra_status_rows.append({
            "skuid": sku,
            "提报状态_最终": merged.final_status,
            "提报结果标签": merged.result_label,
            "提报详情": merged.detail,
            "促销状态_原始": merged.promo_status,
            "提报记录数": merged.record_count,
            "提报记录来源": EXTRA_SOURCE,
            "资料补齐来源": fill_source,
            "处理方式": "已追加到筛选结果，参与后续复提状态链路",
            "说明": "后台提报情况中存在，但不在本轮工具主表；已纳入后续复提状态链路",
        })
    style_sheet(sheet)

    for title in SUMMARY_SHEETS:
        delete_sheet_if_exists(workbook, title)
    append_rows_sheet(workbook, "提报情况原始", status_headers, status_rows)
    append_rows_sheet(workbook, "提报情况归并", ["skuid", "提报状态_最终", "提报结果标签", "提报详情", "促销状态_原始", "提报记录数", "采用记录行号", "采用记录时间"], merged_rows)
    append_rows_sheet(
        workbook,
        "非本轮主表提报记录",
        ["skuid", "提报状态_最终", "提报结果标签", "提报详情", "促销状态_原始", "提报记录数", "提报记录来源", "资料补齐来源", "处理方式", "说明"],
        extra_status_rows,
    )
    append_rows_sheet(
        workbook,
        "业务确认明细",
        ["SKU", "业务确认字段", "业务确认原值", BUSINESS_PARTICIPATION_COLUMN, "说明"],
        existing_detail_rows or confirmation_detail_rows,
    )
    summary_rows = build_summary_rows(final_rows_for_summary, merged_rows)
    append_summary_sheet(workbook, summary_rows)
    participation_counter = Counter(row.get(BUSINESS_PARTICIPATION_COLUMN, "") for row in final_rows_for_summary)

    logs = [
        "提报情况整理开始",
        f"最终结果文件：{final_file}",
        *status_logs,
        *merge_logs,
        *confirm_logs,
        f"筛选结果行数：{len(final_rows_for_summary)}",
        f"匹配提报情况SKU数：{matched}",
        f"未提报/未匹配SKU数：{missing}",
        f"后台额外提报SKU数：{len(extra_skus)}",
        f"已追加到筛选结果的额外SKU数：{len(extra_status_rows)}",
        f"业务确认参与报名SKU数：{participation_counter.get(JOIN_STATUS_YES, 0)}",
        f"业务确认不报名SKU数：{participation_counter.get(JOIN_STATUS_NO, 0)}",
        f"业务确认未确认SKU数：{participation_counter.get(JOIN_STATUS_UNKNOWN, 0)}",
        "规则：同SKU成功优先；详情含严重错误标记为重复提报并视为提报成功；无成功时取最新失败记录。",
    ]
    output_dir = config.output_root / "最终结果"
    output_dir.mkdir(parents=True, exist_ok=True)
    output_file = output_path_for(config, status_file, output_dir)
    logs.append(f"输出文件：{output_file}")
    log_sheet = workbook.create_sheet("提报整理日志")
    for line in logs:
        log_sheet.append([line])
    style_sheet(log_sheet)

    workbook.save(output_file)
    workbook.close()
    return {
        "status_file": status_file,
        "final_file": final_file,
        "confirm_file": confirm_file,
        "output_file": output_file,
        "status_records": len(status_records),
        "merged_skus": len(merged_by_sku),
        "matched_skus": matched,
        "missing_skus": missing,
        "extra_skus": len(extra_skus),
        "appended_extra_skus": len(extra_status_rows),
        "business_join_skus": participation_counter.get(JOIN_STATUS_YES, 0),
        "business_no_join_skus": participation_counter.get(JOIN_STATUS_NO, 0),
        "business_unknown_skus": participation_counter.get(JOIN_STATUS_UNKNOWN, 0),
        "summary_rows": summary_rows,
        "logs": logs,
    }
