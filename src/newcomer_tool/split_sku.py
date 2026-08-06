from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

from .batch import archive_matching_files
from .config import AppConfig


def extract_skus(source: Path, source_sheet: str | None, source_column: str) -> list[str]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    try:
        worksheet = workbook[source_sheet] if source_sheet and source_sheet in workbook.sheetnames else workbook.active
        header_row = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
        sku_index = None
        for candidate in [source_column, "sku", "SKU", "SKU(含影)"]:
            if candidate in header_row:
                sku_index = header_row.index(candidate)
                break
        if sku_index is None:
            raise ValueError(f"未找到SKU字段：{header_row}")
        skus = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            value = row[sku_index]
            if value is not None and str(value).strip():
                skus.append(str(value).strip())
        return skus
    finally:
        workbook.close()


def save_sku_file(path: Path, sheet_name: str, skus: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(["sku"])
    for sku in skus:
        worksheet.append([sku])
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 22
    workbook.save(path)


def latest_filter_result(config: AppConfig) -> Path:
    date_tag = str(config.get("date_tag"))
    sku_file = config.filter_output_dir / f"筛品结果SKU_{date_tag}.xlsx"
    if sku_file.exists():
        return sku_file
    result_file = config.filter_output_dir / f"新人价筛品结果_{date_tag}.xlsx"
    if result_file.exists():
        return result_file
    candidates = sorted(config.filter_output_dir.glob("*.xlsx"), key=lambda item: item.stat().st_mtime, reverse=True)
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f"未找到筛品结果或SKU文件：{sku_file}")


def run_split(config: AppConfig, source: Path | None = None) -> dict[str, Any]:
    source_path = source or latest_filter_result(config)
    output_dir = config.split_output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    skus = extract_skus(source_path, str(config.get("split_source_sheet", "汇总")), str(config.get("split_source_column", "sku")))
    base_name = str(config.get("split_base_name", "8月第一批查价前sku"))
    chunk_size = int(config.get("split_chunk_size", 5000))
    if chunk_size <= 0:
        raise ValueError("split_chunk_size 必须大于 0")
    archive_result = archive_matching_files(
        output_dir,
        [f"{base_name}.xlsx", f"{base_name}_PART*.xlsx"],
        "split_sku",
    )
    summary_path = output_dir / f"{base_name}.xlsx"
    save_sku_file(summary_path, "汇总", skus)
    part_paths = []
    for start in range(0, len(skus), chunk_size):
        part_no = start // chunk_size + 1
        part_path = output_dir / f"{base_name}_PART{part_no:02d}.xlsx"
        save_sku_file(part_path, f"PART{part_no:02d}", skus[start:start + chunk_size])
        part_paths.append(part_path)
    return {"source": source_path, "summary": summary_path, "parts": part_paths, "sku_count": len(skus), **archive_result}
