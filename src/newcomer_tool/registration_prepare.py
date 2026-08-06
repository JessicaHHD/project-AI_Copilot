from __future__ import annotations

from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from .batch import archive_matching_files, registration_submit_base, unique_path
from .config import AppConfig


BUSINESS_PARTICIPATION_COLUMN = "业务是否参与报名"
BUSINESS_DETAIL_SHEET = "业务确认明细"
JOIN_STATUS_YES = "参与报名"
JOIN_STATUS_NO = "不报名"
JOIN_STATUS_UNKNOWN = "未确认"
NO_VALUES = {"否", "n", "no", "不", "不参加", "不报名", "不提报", "退出", "取消", "放弃", "false", "0"}
EXIT_VALUES = {"是", "y", "yes", "true", "1", "退出", "不参加", "不报名", "不提报", "取消", "放弃"}


def normalize(value: object) -> str:
    return str(value or "").strip().replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")


def to_price(value: object) -> Decimal | None:
    text = str(value or "").strip().replace(",", "")
    if not text or text in {"-", "--"}:
        return None
    try:
        return Decimal(text).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    except InvalidOperation:
        return None


def format_price(value: Decimal) -> float:
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def text(value: object) -> str:
    return str(value or "").strip()


def sku_text(value: object) -> str:
    raw = text(value)
    if raw.endswith(".0"):
        raw = raw[:-2]
    return raw


def find_index(headers: list[object], candidates: list[str]) -> int | None:
    normalized = [normalize(header).lower() for header in headers]
    for candidate in candidates:
        target = normalize(candidate).lower()
        if target in normalized:
            return normalized.index(target)
    for index, header in enumerate(normalized):
        if any(normalize(candidate).lower() in header for candidate in candidates):
            return index
    return None


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, list[object], int, int, int, str]:
    for row_index, row in enumerate(rows[:30]):
        headers = list(row)
        sku_index = find_index(headers, ["sku", "skuid", "SKU", "SKU(含影)"])
        price_index = find_index(headers, ["期望新人价", "新人价", "首单新人价", "促销价"])
        join_index = find_index(headers, ["是否退出", "是否登记", "是否报名", "是否参加", "报名", "是否提报"])
        if sku_index is not None and price_index is not None and join_index is not None:
            return row_index, headers, sku_index, price_index, join_index, str(headers[join_index] or "")
    raise ValueError("未找到表头行：需要同时包含 SKU、期望新人价/新人价、是否退出/是否报名 字段")


def is_not_joining(value: object, field_name: str) -> bool:
    text = normalize(value).lower()
    if not text:
        return False
    if "退出" in normalize(field_name):
        if text in EXIT_VALUES:
            return True
        return any(keyword in text for keyword in ["不参加", "不报名", "不提报", "退出", "取消", "放弃"])
    if text in NO_VALUES:
        return True
    return any(keyword in text for keyword in ["不参加", "不报名", "不提报", "退出", "取消", "放弃"])


def business_participation(field_name: str, raw_value: Any, found: bool = True) -> tuple[str, str]:
    if not found:
        return JOIN_STATUS_UNKNOWN, "业务确认表未找到该SKU"
    value = normalize(raw_value).lower()
    if not value:
        return JOIN_STATUS_YES, "空白默认参与报名"
    field = normalize(field_name)
    if "退出" in field:
        if value in EXIT_VALUES or any(keyword in value for keyword in ["不参加", "不报名", "不提报", "退出", "取消", "放弃"]):
            return JOIN_STATUS_NO, "业务确认字段表示退出/不报名"
        return JOIN_STATUS_YES, "未命中退出/不报名关键词"
    if value in NO_VALUES or any(keyword in value for keyword in ["不参加", "不报名", "不提报", "退出", "取消", "放弃"]):
        return JOIN_STATUS_NO, "业务确认字段表示不报名"
    return JOIN_STATUS_YES, "未命中不报名关键词"


def read_confirmation_rows(path: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, str]], list[dict[str, str]], list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        if hasattr(worksheet, "reset_dimensions"):
            worksheet.reset_dimensions()
        all_rows = list(worksheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    header_row, headers, sku_index, price_index, join_index, join_field = find_header_row(all_rows)
    selected: list[dict[str, Any]] = []
    confirm_map: dict[str, dict[str, str]] = {}
    detail_rows: list[dict[str, str]] = []
    skipped_no = 0
    skipped_missing = 0
    seen_skus: set[str] = set()
    duplicate_skus = 0

    for row in all_rows[header_row + 1:]:
        join_value = row[join_index] if join_index < len(row) else ""
        sku = sku_text(row[sku_index] if sku_index < len(row) else "")
        if not sku:
            continue
        status, note = business_participation(join_field, join_value)
        item = {
            "SKU": sku,
            "业务确认字段": join_field,
            "业务确认原值": text(join_value),
            BUSINESS_PARTICIPATION_COLUMN: status,
            "说明": note,
        }
        confirm_map[sku] = item
        detail_rows.append(item)
        if status == JOIN_STATUS_NO:
            skipped_no += 1
            continue
        price = to_price(row[price_index] if price_index < len(row) else "")
        if not sku or price is None:
            skipped_missing += 1
            continue
        if sku in seen_skus:
            duplicate_skus += 1
            continue
        seen_skus.add(sku)
        selected.append({"skuid": sku, "促销价": price, BUSINESS_PARTICIPATION_COLUMN: status})

    logs = [
        f"确认表文件：{path}",
        f"识别表头行：第 {header_row + 1} 行",
        f"识别报名判断字段：{join_field}",
        f"采销确认SKU数：{len(confirm_map)}",
        f"报名确认行数：{len(selected)}",
        f"明确退出/不报名跳过：{skipped_no}",
        f"SKU或新人价缺失跳过：{skipped_missing}",
        f"重复SKU跳过：{duplicate_skus}",
    ]
    return selected, confirm_map, detail_rows, logs


def write_submit_file(path: Path, rows: list[dict[str, Any]], sheet_name: str, include_business_status: bool = False) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    headers = ["skuid", "促销价"]
    if include_business_status:
        headers.append(BUSINESS_PARTICIPATION_COLUMN)
    worksheet.append(headers)
    for row in rows:
        values = [row["skuid"], format_price(row["促销价"])]
        if include_business_status:
            values.append(row.get(BUSINESS_PARTICIPATION_COLUMN, ""))
        worksheet.append(values)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 22
    worksheet.column_dimensions["B"].width = 12
    workbook.save(path)


def latest_final_pricing_file(config: AppConfig) -> Path:
    output_dir = config.output_root / "最终结果"
    if not output_dir.exists():
        raise FileNotFoundError(f"最终结果目录不存在：{output_dir}")
    candidates = sorted(
        (
            path for path in output_dir.glob("*筛品查价表*.xlsx")
            if path.is_file()
            and not path.name.startswith("~$")
            and "提报情况整理表" not in path.name
        ),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"未找到筛品查价最终表：{output_dir / '*筛品查价表*.xlsx'}")
    return candidates[0]


def delete_sheet_if_exists(workbook, title: str) -> None:
    if title in workbook.sheetnames:
        del workbook[title]


def style_header(worksheet) -> None:
    if worksheet.max_row < 1:
        return
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        header = column_cells[0]
        worksheet.column_dimensions[header.column_letter].width = min(max(len(text(header.value)) + 2, 12), 40)


def write_business_detail_sheet(workbook, detail_rows: list[dict[str, str]]) -> None:
    headers = ["SKU", "业务确认字段", "业务确认原值", BUSINESS_PARTICIPATION_COLUMN, "说明"]
    delete_sheet_if_exists(workbook, BUSINESS_DETAIL_SHEET)
    worksheet = workbook.create_sheet(BUSINESS_DETAIL_SHEET)
    worksheet.append(headers)
    for row in detail_rows:
        worksheet.append([row.get(header, "") for header in headers])
    style_header(worksheet)


def update_final_pricing_business_confirmation(
    config: AppConfig,
    confirm_map: dict[str, dict[str, str]],
    detail_rows: list[dict[str, str]],
) -> dict[str, Any]:
    final_file = latest_final_pricing_file(config)
    workbook = load_workbook(final_file)
    try:
        if "筛选结果" not in workbook.sheetnames:
            raise ValueError(f"筛品查价最终表缺少“筛选结果”sheet：{final_file}")
        worksheet = workbook["筛选结果"]
        headers = [text(worksheet.cell(1, column).value) for column in range(1, worksheet.max_column + 1)]
        sku_index = find_index(headers, [str(config.get("sku_column")), "SKU(含影)", "skuid", "skuId", "SKU", "sku"])
        if sku_index is None:
            raise ValueError(f"筛选结果缺少SKU字段：{final_file}")
        business_index = find_index(headers, [BUSINESS_PARTICIPATION_COLUMN])
        if business_index is None:
            business_index = len(headers)
            worksheet.cell(1, business_index + 1, BUSINESS_PARTICIPATION_COLUMN)
            worksheet.cell(1, business_index + 1).font = Font(bold=True)

        final_detail_rows = list(detail_rows)
        join_count = 0
        no_join_count = 0
        unknown_count = 0
        for row_index in range(2, worksheet.max_row + 1):
            sku = sku_text(worksheet.cell(row_index, sku_index + 1).value)
            if not sku:
                continue
            item = confirm_map.get(sku)
            if item:
                status = item.get(BUSINESS_PARTICIPATION_COLUMN, JOIN_STATUS_UNKNOWN)
            else:
                status, note = business_participation("", "", found=False)
                final_detail_rows.append({
                    "SKU": sku,
                    "业务确认字段": "",
                    "业务确认原值": "",
                    BUSINESS_PARTICIPATION_COLUMN: status,
                    "说明": note,
                })
            if status == JOIN_STATUS_YES:
                join_count += 1
            elif status == JOIN_STATUS_NO:
                no_join_count += 1
            else:
                unknown_count += 1
            worksheet.cell(row_index, business_index + 1, status)

        write_business_detail_sheet(workbook, final_detail_rows)
        workbook.save(final_file)
    finally:
        workbook.close()
    return {
        "final_file": final_file,
        "business_join_skus": join_count,
        "business_no_join_skus": no_join_count,
        "business_unknown_skus": unknown_count,
        "confirmation_detail_rows": final_detail_rows,
    }


def write_log_file(path: Path, logs: list[str]) -> None:
    path.write_text("\n".join(logs) + "\n", encoding="utf-8")


def run_registration_prepare(config: AppConfig, source_file: Path) -> dict[str, Any]:
    if not source_file.exists():
        raise FileNotFoundError(f"采销确认表不存在：{source_file}")
    rows, confirm_map, detail_rows, logs = read_confirmation_rows(source_file)
    final_update = update_final_pricing_business_confirmation(config, confirm_map, detail_rows)
    if not rows:
        raise ValueError(f"未提取到任何可提报SKU，已更新筛品查价最终表：{final_update['final_file']}。请检查 SKU、期望新人价/新人价、是否退出 字段。")

    output_dir = Path(str(config.get("registration_submit_dir")))
    output_dir.mkdir(parents=True, exist_ok=True)
    base_name = registration_submit_base(config)
    archive_result = archive_matching_files(
        output_dir,
        [f"{base_name}.xlsx", f"{base_name}_PART*.xlsx"],
        "registration_submit",
    )
    chunk_size = int(config.get("registration_chunk_size", 5000) or 5000)
    if chunk_size <= 0:
        raise ValueError("registration_chunk_size 必须大于 0")

    summary_path = unique_path(output_dir / f"{base_name}.xlsx")
    write_submit_file(summary_path, rows, "汇总", include_business_status=True)
    part_paths: list[Path] = []
    for start in range(0, len(rows), chunk_size):
        part_no = start // chunk_size + 1
        part_path = unique_path(output_dir / f"{base_name}_PART{part_no:02d}.xlsx")
        write_submit_file(part_path, rows[start:start + chunk_size], f"PART{part_no:02d}")
        part_paths.append(part_path)
    logs.extend([
        f"已更新筛品查价最终表：{final_update['final_file']}",
        f"业务确认参与报名SKU数：{final_update['business_join_skus']}",
        f"业务确认不报名SKU数：{final_update['business_no_join_skus']}",
        f"业务确认未确认SKU数：{final_update['business_unknown_skus']}",
        f"已归档旧提报文件数：{len(archive_result['archived_files'])}",
        f"旧提报文件归档目录：{archive_result['archive_dir'] or '无'}",
        f"输出汇总文件：{summary_path}",
        f"拆分大小：{chunk_size}",
        f"输出PART文件数：{len(part_paths)}",
    ])
    log_path = unique_path(output_dir / f"{base_name}_生成日志.txt")
    write_log_file(log_path, logs)
    return {
        "source": source_file,
        "summary": summary_path,
        "parts": part_paths,
        "log": log_path,
        "rows": len(rows),
        **archive_result,
        **final_update,
        "logs": logs,
    }
