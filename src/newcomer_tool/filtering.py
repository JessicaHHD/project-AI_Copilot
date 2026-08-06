from __future__ import annotations

import csv
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .config import AppConfig


def detect_encoding(path: Path) -> str:
    for encoding in ("utf-8-sig", "gb18030", "gbk", "cp936", "utf-16"):
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                file.read(8192)
            return encoding
        except UnicodeError:
            continue
    return "gb18030"


def read_table(path: Path) -> tuple[list[str], list[dict[str, str]], str]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            worksheet = workbook.active
            rows_iter = worksheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(rows_iter)]
            rows = []
            for row in rows_iter:
                rows.append({headers[index]: str(value or "") for index, value in enumerate(row[: len(headers)])})
            return headers, rows, "excel"
        finally:
            workbook.close()
    encoding = detect_encoding(path)
    with path.open("r", encoding=encoding, newline="") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        rows = [{key: (value or "") for key, value in row.items()} for row in reader]
    return headers, rows, encoding


def normalize(value: str) -> str:
    return value.replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")


def variants(name: str) -> list[str]:
    names = [name]
    if "_" in name:
        names.append(name.replace("_", "*"))
    if "*" in name:
        names.append(name.replace("*", "_"))
    return [item for item in dict.fromkeys(names) if item]


def find_column(headers: list[str], candidates: list[str], required: bool = True) -> str | None:
    expanded = []
    for candidate in candidates:
        expanded.extend(variants(str(candidate or "")))
    expanded = [item for item in expanded if item]
    normalized = {normalize(header): header for header in headers}
    for candidate in expanded:
        if normalize(candidate) in normalized:
            return normalized[normalize(candidate)]
    for header in headers:
        compact = normalize(header)
        if any(normalize(candidate) in compact for candidate in expanded):
            return header
    if required:
        raise KeyError("缺少必要字段：" + "、".join(expanded))
    return None


def to_decimal(value: object) -> Decimal:
    text = str(value or "").strip().replace(",", "")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def format_decimal(value: object) -> str:
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return str(int(value))
        text = format(value.normalize(), "f")
        return text.rstrip("0").rstrip(".") if "." in text else text
    return str(value or "")


def add_audit(rows: list[dict[str, str]], item: str, status: str, detail: str) -> None:
    rows.append({"检查项": item, "状态": status, "说明": detail})


def audit_status(rows: list[dict[str, str]]) -> str:
    statuses = {row["状态"] for row in rows}
    if "失败" in statuses:
        return "FAIL"
    if "预警" in statuses:
        return "WARN"
    return "PASS"


def load_sku_set(path_text: str, sku_column_name: str) -> set[str]:
    if not path_text:
        return set()
    path = Path(path_text)
    if not path.exists():
        raise FileNotFoundError(f"SKU名单文件不存在：{path}")
    headers, rows, _ = read_table(path)
    sku_column = find_column(headers, [sku_column_name, "sku", "SKU", "SKU(含影)"], required=False) or headers[0]
    return {str(row.get(sku_column, "")).strip() for row in rows if str(row.get(sku_column, "")).strip()}


def dedupe(rows: list[dict[str, str]], sku_column: str, metric_columns: list[str]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in rows:
        sku = str(row.get(sku_column, "")).strip()
        if sku not in grouped:
            grouped[sku] = dict(row)
            for column in metric_columns:
                grouped[sku][column] = to_decimal(row.get(column, ""))
        else:
            for column in metric_columns:
                grouped[sku][column] = to_decimal(grouped[sku].get(column, "")) + to_decimal(row.get(column, ""))
    return list(grouped.values())


def remove_skus(rows: list[dict[str, Any]], sku_column: str, skus: set[str]) -> tuple[list[dict[str, Any]], int]:
    if not skus:
        return rows, 0
    kept = [row for row in rows if str(row.get(sku_column, "")).strip() not in skus]
    return kept, len(rows) - len(kept)


def sku_text(value: object) -> str:
    raw = str(value or "").strip()
    if raw.endswith(".0"):
        raw = raw[:-2]
    return raw


def config_list(config: AppConfig, key: str) -> list[str]:
    value = config.get(key, [])
    if isinstance(value, str):
        return [value] if value else []
    return [str(item) for item in (value or []) if str(item)]


def contains_keyword(value: str, keywords: list[str]) -> bool:
    return any(keyword and keyword in value for keyword in keywords)


def target_month_text(config: AppConfig) -> str:
    threshold = int(config.get("registration_month_threshold_day", 25) or 25)
    now = datetime.now()
    year = now.year
    month = now.month
    if now.day >= threshold:
        month += 1
        if month > 12:
            month = 1
            year += 1
    return f"{year}年{month}月"


def latest_resubmit_final_file(config: AppConfig) -> Path | None:
    configured = str(config.get("resubmit_final_file", "") or "").strip()
    if configured:
        path = Path(configured)
        if not path.exists():
            raise FileNotFoundError(f"复提最终结果文件不存在：{path}")
        return path
    directory = config.output_root / "最终结果"
    if not directory.exists():
        return None
    patterns = ["*提报情况整理表*.xlsx", "新人价提报整理_*.xlsx"]
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(path for path in directory.glob(pattern) if path.is_file() and not path.name.startswith("~$"))
    candidates = sorted(set(candidates), key=lambda item: item.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def source_auto_candidates(config: AppConfig) -> list[Path]:
    directory = Path(str(config.get("source_auto_dir", "") or Path.home() / "Downloads"))
    if not directory.exists():
        return []
    patterns = config_list(config, "source_auto_patterns") or ["*近30天动销大于200*.csv", "*近30天动销大于200*.xlsx", "*近30天动销大于200*.xlsm"]
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(path for path in directory.glob(pattern) if path.is_file() and not path.name.startswith("~$"))
    return sorted(set(candidates), key=lambda item: item.stat().st_mtime, reverse=True)


def resolve_source_file(config: AppConfig, resubmit_pool: dict[str, Any] | None) -> tuple[Path | None, str]:
    manual_text = str(config.get("source_file", "") or "").strip()
    manual_source = Path(manual_text) if manual_text else None
    manual_exists = bool(manual_source and manual_source.exists() and manual_source.is_file())
    auto_detect = bool(config.get("source_auto_detect", True))
    reference_file = resubmit_pool["path"] if resubmit_pool else None
    reference_mtime = reference_file.stat().st_mtime if reference_file else None

    if auto_detect:
        candidates = source_auto_candidates(config)
        latest_auto = candidates[0] if candidates else None
        if latest_auto and (reference_mtime is None or latest_auto.stat().st_mtime > reference_mtime):
            return latest_auto, f"自动识别到更新的大盘源表：{latest_auto}"
        if manual_exists and (reference_mtime is None or manual_source.stat().st_mtime > reference_mtime):
            return manual_source, f"使用手动配置的大盘源表：{manual_source}"
        if reference_file:
            latest_text = f"；下载目录最新候选：{latest_auto}" if latest_auto else "；下载目录未找到候选源表"
            return None, f"未发现晚于最近提报整理结果的新增大盘源表，将直接走复提模式{latest_text}"
        if manual_exists:
            return manual_source, f"初次流程使用手动配置的大盘源表：{manual_source}"
        return None, "未配置源表，且下载目录未自动识别到大盘源表"

    if manual_exists:
        return manual_source, f"自动识别关闭，使用手动配置的大盘源表：{manual_source}"
    return None, f"自动识别关闭，且手动源表不存在：{manual_text or '(未配置)'}"


def read_excel_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"文件缺少“{sheet_name}”sheet：{path}")
        worksheet = workbook[sheet_name]
        if hasattr(worksheet, "reset_dimensions"):
            worksheet.reset_dimensions()
        rows_iter = worksheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(rows_iter)]
        except StopIteration:
            return [], []
        rows: list[dict[str, Any]] = []
        for row in rows_iter:
            rows.append({headers[index]: row[index] if index < len(row) else None for index in range(len(headers))})
        return headers, rows
    finally:
        workbook.close()


def load_resubmit_pool(config: AppConfig, sku_column_name: str) -> dict[str, Any] | None:
    path = latest_resubmit_final_file(config)
    if path is None:
        return None
    headers, rows = read_excel_sheet(path, "筛选结果")
    if not headers:
        raise ValueError(f"复提最终结果文件“筛选结果”为空：{path}")
    sku_col = find_column(headers, [sku_column_name, "SKU(含影)", "skuid", "skuId", "SKU", "sku"])
    status_col = find_column(headers, ["提报状态_最终", "提报状态"], required=False)
    label_col = find_column(headers, ["提报结果标签", "提报失败标签", "促销状态"], required=False)
    detail_col = find_column(headers, ["提报详情", "详情", "失败原因"], required=False)
    business_join_col = find_column(headers, ["业务是否参与报名"], required=False)
    exit_col = find_column(headers, ["是否退出_采销确认", "是否退出", "是否报名", "是否参加"], required=False)
    missing = []
    if not status_col:
        missing.append("提报状态_最终")
    if not label_col:
        missing.append("提报结果标签")
    if not detail_col:
        missing.append("提报详情")
    if missing:
        raise ValueError(f"复提最终结果文件缺少字段：{'、'.join(missing)}；文件：{path}")

    success_keywords = config_list(config, "resubmit_success_keywords")
    exit_keywords = config_list(config, "resubmit_exit_keywords")
    row_by_sku: dict[str, dict[str, Any]] = {}
    resubmit_skus: set[str] = set()
    failed_resubmit_skus: set[str] = set()
    success_resubmit_skus: set[str] = set()
    ordinary_success_skus: set[str] = set()
    exit_skus: set[str] = set()
    missing_status_skus: set[str] = set()
    order: list[str] = []

    for row in rows:
        sku = sku_text(row.get(sku_col, ""))
        if not sku:
            continue
        if sku not in row_by_sku:
            order.append(sku)
        current_row = dict(row)
        current_row[sku_col] = sku
        row_by_sku[sku] = current_row

        business_join_value = str(row.get(business_join_col, "") or "") if business_join_col else ""
        exit_value = str(row.get(exit_col, "") or "") if exit_col else ""
        if "不报名" in business_join_value or contains_keyword(exit_value, exit_keywords):
            exit_skus.add(sku)
            resubmit_skus.discard(sku)
            failed_resubmit_skus.discard(sku)
            success_resubmit_skus.discard(sku)
            ordinary_success_skus.discard(sku)
            continue

        status = str(row.get(status_col, "") or "")
        label = str(row.get(label_col, "") or "") if label_col else ""
        detail = str(row.get(detail_col, "") or "") if detail_col else ""
        combined = f"{label} {detail}"
        if "提报失败" in status:
            resubmit_skus.add(sku)
            failed_resubmit_skus.add(sku)
            ordinary_success_skus.discard(sku)
        elif "提报成功" in status:
            if contains_keyword(combined, success_keywords):
                resubmit_skus.add(sku)
                success_resubmit_skus.add(sku)
                ordinary_success_skus.discard(sku)
            else:
                if sku not in resubmit_skus:
                    ordinary_success_skus.add(sku)
        elif "未提报" in status or "未匹配" in status:
            missing_status_skus.add(sku)

    return {
        "path": path,
        "headers": headers,
        "rows": rows,
        "sku_col": sku_col,
        "row_by_sku": row_by_sku,
        "order": order,
        "resubmit_skus": resubmit_skus,
        "failed_resubmit_skus": failed_resubmit_skus,
        "success_resubmit_skus": success_resubmit_skus,
        "ordinary_success_skus": ordinary_success_skus,
        "exit_skus": exit_skus,
        "missing_status_skus": missing_status_skus,
    }


def append_resubmit_rows(rows: list[dict[str, Any]], sku_column: str, resubmit_pool: dict[str, Any]) -> tuple[list[dict[str, Any]], int]:
    existing = {sku_text(row.get(sku_column, "")) for row in rows if sku_text(row.get(sku_column, ""))}
    result = list(rows)
    added = 0
    for sku in resubmit_pool["order"]:
        if sku not in resubmit_pool["resubmit_skus"] or sku in existing:
            continue
        row = dict(resubmit_pool["row_by_sku"].get(sku, {}))
        row[sku_column] = sku
        result.append(row)
        existing.add(sku)
        added += 1
    return result, added


def write_filter_outputs(
    config: AppConfig,
    output_dir: Path,
    date_tag: str,
    headers: list[str],
    final_rows: list[dict[str, Any]],
    logs: list[str],
    audit: list[dict[str, str]],
    sku_col: str,
    source: Path | None,
    source_rows_count: int,
    source_type: str,
) -> dict[str, Any]:
    add_audit(audit, "最终结果数量", "通过" if final_rows else "失败", f"最终剩余 {len(final_rows)} 个")
    skus = [sku_text(row.get(sku_col, "")) for row in final_rows if sku_text(row.get(sku_col, ""))]
    sku_file = output_dir / f"筛品结果SKU_{date_tag}.xlsx"
    write_sku_file(sku_file, skus)
    add_audit(audit, "下一模块SKU文件", "通过" if len(skus) == len(final_rows) else "预警", f"已生成 {len(skus)} 条：{sku_file}")

    xlsx = output_dir / f"新人价筛品结果_{date_tag}.xlsx"
    log_path = output_dir / f"新人价筛品日志_{date_tag}.txt"
    add_audit(audit, "输出文件检查", "通过", "已生成Excel、日志和SKU文件")
    write_result_xlsx(xlsx, headers, final_rows, logs)
    log_path.write_text("\n".join(logs) + "\n", encoding="utf-8")

    audit_report = output_dir / f"筛品过程记录_{date_tag}.txt"
    audit_lines = ["检查项\t状态\t说明"]
    audit_lines.extend([f"{row['检查项']}\t{row['状态']}\t{row['说明']}" for row in audit])
    process_lines = [
        "【筛品日志】",
        *logs,
        "",
        "【审核结果】",
        *audit_lines,
    ]
    audit_report.write_text("\n".join(process_lines) + "\n", encoding="utf-8")

    csv_path = None
    if bool(config.get("export_csv", False)):
        csv_path = output_dir / f"新人价筛品结果_{date_tag}.csv"
        write_csv(csv_path, headers, final_rows)

    return {
        "status": audit_status(audit),
        "warnings": [row for row in audit if row["状态"] == "预警"],
        "failures": [row for row in audit if row["状态"] == "失败"],
        "source_file": source,
        "source_rows": source_rows_count,
        "rows": len(final_rows),
        "xlsx": xlsx,
        "log": log_path,
        "sku_file": sku_file,
        "audit_report": audit_report,
        "csv": csv_path,
        "logs": logs,
        "audit_rows": audit,
        "source_type": source_type,
    }


def write_sku_file(path: Path, skus: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "汇总"
    worksheet.append(["sku"])
    for sku in skus:
        worksheet.append([sku])
    worksheet.freeze_panes = "A2"
    worksheet.column_dimensions["A"].width = 22
    workbook.save(path)


def write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({header: format_decimal(row.get(header, "")) for header in headers})



def write_result_xlsx(path: Path, headers: list[str], rows: list[dict[str, Any]], logs: list[str]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "筛选结果"
    sheet.append(headers)
    for row in rows:
        sheet.append([format_decimal(row.get(header, "")) for header in headers])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(column[0].value or "")) + 2, 12), 28)

    log_sheet = workbook.create_sheet("筛品日志")
    for line in logs:
        log_sheet.append([line])
    log_sheet.column_dimensions["A"].width = 120
    for row in log_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    workbook.save(path)

def validate_columns(headers: list[str], config: AppConfig) -> tuple[dict[str, str], list[str]]:
    specs = {
        "sku": [config.get("sku_column"), "SKU(含影)", "sku", "SKU"],
        "name": [config.get("product_name_column"), "sku名称", "商品名称"],
        "mapping": [config.get("mapping_column"), "1自营3POP_映射"],
        "shop": [config.get("shop_column"), "店铺名称"],
        "department": [config.get("department_column"), "1级部门名称（类目运营）", "一级部门"],
        "category2": [config.get("category2_column"), "商品二级分类名称"],
        "erp": [config.get("erp_column"), "销售员ERP帐号", "销售员ERP账号"],
    }
    found: dict[str, str] = {}
    missing: list[str] = []
    for key, candidates in specs.items():
        column = find_column(headers, [str(item) for item in candidates], required=False)
        if column:
            found[key] = column
        else:
            missing.append("/".join([str(item) for item in candidates if item]))
    metric_columns = []
    for metric in config.get("metric_sum_columns", []):
        column = find_column(headers, [str(metric)], required=False)
        if column:
            metric_columns.append(column)
        else:
            missing.append(str(metric))
    found["metrics"] = "||".join(metric_columns)
    return found, missing


def run_filtering(config: AppConfig) -> dict[str, Any]:
    output_dir = config.filter_output_dir
    output_dir.mkdir(parents=True, exist_ok=True)
    date_tag = str(config.get("date_tag"))
    audit: list[dict[str, str]] = []
    logs: list[str] = []
    target_month = target_month_text(config)
    mode = str(config.get("filter_round_mode", "auto") or "auto").strip().lower()
    if mode not in {"auto", "source_first", "resubmit_only"}:
        raise ValueError("filter_round_mode 只能是 auto/source_first/resubmit_only")

    resubmit_pool = load_resubmit_pool(config, str(config.get("sku_column")))
    source, source_note = resolve_source_file(config, resubmit_pool)
    source_exists = bool(source and source.exists() and source.is_file())
    if mode == "resubmit_only":
        active_mode = "resubmit_only"
    elif mode == "source_first":
        active_mode = "source_first"
    else:
        active_mode = "source_first" if source_exists else "resubmit_only" if resubmit_pool else "source_first"

    logs.append(f"0、筛品运行模式：{active_mode}；当前目标新人价计划月份：{target_month}。")
    logs.append(f"0、源表判断：{source_note}。")
    if resubmit_pool:
        logs.append(f"0、已读取最近提报整理最终表：{resubmit_pool['path']}。")
        add_audit(audit, "提报整理最终表", "通过", f"已读取：{resubmit_pool['path']}")
    else:
        logs.append("0、未读取到提报情况整理表；如为初次提报，这是正常情况，将按原筛品流程执行。")
        add_audit(audit, "提报情况整理表", "通过", "未找到提报情况整理表；按初次提报/原筛品流程执行")

    if active_mode == "resubmit_only":
        if not resubmit_pool:
            add_audit(audit, "复提候选来源", "失败", "无源货盘且未找到提报整理最终表")
            raise FileNotFoundError("无源货盘且未找到提报整理最终表，无法生成复提候选")
        headers = list(resubmit_pool["headers"])
        sku_col = str(resubmit_pool["sku_col"])
        rows = [
            dict(resubmit_pool["row_by_sku"][sku])
            for sku in resubmit_pool["order"]
            if sku in resubmit_pool["resubmit_skus"]
        ]
        logs.append(
            "1、无新的大盘订单源表，直接从最近提报整理最终表生成复提候选："
            f"提报失败{len(resubmit_pool['failed_resubmit_skus'])}个，"
            f"成功但需复提{len(resubmit_pool['success_resubmit_skus'])}个，"
            f"采销退出剔除{len(resubmit_pool['exit_skus'])}个，"
            f"普通成功不复提{len(resubmit_pool['ordinary_success_skus'])}个。"
        )
        add_audit(audit, "复提候选生成", "通过" if rows else "失败", f"复提候选 {len(rows)} 个")
        erp_col = find_column(headers, [str(config.get("erp_column")), "销售员ERP帐号", "销售员ERP账号"], required=False)
        if erp_col:
            excluded_erps = set(config.get("exclude_erp_accounts", []))
            removed_erp = [row for row in rows if str(row.get(erp_col, "")).strip() in excluded_erps]
            final_rows = [row for row in rows if str(row.get(erp_col, "")).strip() not in excluded_erps]
            logs.append(f"2、剔除指定ERP帐号（{'、'.join(excluded_erps)}）{len(removed_erp)}个，剩余{len(final_rows)}个。")
        else:
            final_rows = rows
            logs.append("2、未识别ERP字段，跳过ERP帐号剔除。")
            add_audit(audit, "ERP字段识别", "预警", "复提最终表中未识别ERP字段")
        return write_filter_outputs(config, output_dir, date_tag, headers, final_rows, logs, audit, sku_col, resubmit_pool["path"], len(resubmit_pool["rows"]), "resubmit_final")

    if not source_exists:
        add_audit(audit, "源文件存在", "失败", source_note)
        raise FileNotFoundError(source_note)
    assert source is not None
    add_audit(audit, "源文件存在", "通过", f"源文件路径：{source}")

    headers, source_rows, source_type = read_table(source)
    add_audit(audit, "源文件读取", "通过", f"已读取 {len(source_rows)} 条，格式：{source_type}")
    columns, missing = validate_columns(headers, config)
    if missing:
        add_audit(audit, "必填字段完整", "失败", "缺少字段：" + "、".join(missing))
        fail_report = output_dir / f"筛品失败过程记录_{date_tag}.txt"
        audit_lines = ["检查项\t状态\t说明"]
        audit_lines.extend([f"{row['检查项']}\t{row['状态']}\t{row['说明']}" for row in audit])
        process_lines = ["【筛品日志】", *logs, "", "【审核结果】", *audit_lines]
        fail_report.write_text("\n".join(process_lines) + "\n", encoding="utf-8")
        raise KeyError(f"关键字段缺失，已终止：{', '.join(missing)}；过程记录：{fail_report}")
    add_audit(audit, "必填字段完整", "通过", "所有关键字段存在")

    sku_col = columns["sku"]
    name_col = columns["name"]
    mapping_col = columns["mapping"]
    shop_col = columns["shop"]
    dept_col = columns["department"]
    category_col = columns["category2"]
    erp_col = columns["erp"]
    metric_cols = columns["metrics"].split("||")

    targets = list(config.get("target_departments", []))
    step1 = [row for row in source_rows if any(keyword in str(row.get(dept_col, "")) for keyword in targets)]
    logs.append(f"1、取数模型提取近30天动销>200，一级部门为“{'、'.join(targets)}”的sku获得{len(step1)}个sku记录（未去重）。")

    keep_values = set(config.get("keep_mapping_values", []))
    pop_keyword = str(config.get("keep_pop_shop_keyword", ""))
    keep_rows = [row for row in step1 if str(row.get(mapping_col, "")).strip() in keep_values]
    pop_rows = [row for row in step1 if str(row.get(mapping_col, "")).strip().upper() == "POP" and pop_keyword in str(row.get(shop_col, ""))]
    step2 = [row for row in step1 if str(row.get(mapping_col, "")).strip() in keep_values or (str(row.get(mapping_col, "")).strip().upper() == "POP" and pop_keyword in str(row.get(shop_col, "")))]
    step2_removed = len(step1) - len(step2)
    logs.append(f"2、保留{'+'.join(keep_values)}（{len(keep_rows)}）、{pop_keyword}（{len(pop_rows)}），剔除{step2_removed}个，剩余{len(step2)}个。")
    if step1 and step2_removed / len(step1) > 0.6:
        add_audit(audit, "第2步剔除比例", "预警", f"第2步剔除比例 {step2_removed / len(step1):.1%}，请确认自营/京喜规则")
    else:
        add_audit(audit, "第2步剔除比例", "通过", f"第2步剔除 {step2_removed} 个")

    excluded_categories = set(config.get("exclude_category2", []))
    counts = Counter(str(row.get(category_col, "")).strip() for row in step2)
    step3 = [row for row in step2 if str(row.get(category_col, "")).strip() not in excluded_categories]
    category_text = "、".join([f"“{category}”（{counts[category]}个）" for category in config.get("exclude_category2", [])])
    logs.append(f"3、剔除{category_text}，剩余{len(step3)}个。")

    step4 = dedupe(step3, sku_col, metric_cols)
    logs.append(f"4、sku去重后{len(step4)}个，并对“{'、'.join(metric_cols)}”按sku求和。")
    add_audit(audit, "SKU去重检查", "通过", f"去重前 {len(step3)}，去重后 {len(step4)}")

    keywords = list(config.get("exclude_name_keywords", []))
    keyword_counts = {keyword: sum(1 for row in step4 if keyword in str(row.get(name_col, ""))) for keyword in keywords}
    step5 = [row for row in step4 if not any(keyword in str(row.get(name_col, "")) for keyword in keywords)]
    keyword_text = "、".join([f"“{keyword}”（{count}个）" for keyword, count in keyword_counts.items()])
    logs.append(f"5、删除{keyword_text}，实际剔除{len(step4) - len(step5)}个，剩余{len(step5)}个。")

    if resubmit_pool:
        after_success, removed_success = remove_skus(step5, sku_col, resubmit_pool["ordinary_success_skus"])
        after_exit, removed_exit = remove_skus(after_success, sku_col, resubmit_pool["exit_skus"])
        step6, added_resubmit = append_resubmit_rows(after_exit, sku_col, resubmit_pool)
        logs.append(
            "6、根据最近提报整理结果处理当前月复提："
            f"普通成功已在品池剔除{removed_success}个，"
            f"采销退出剔除{removed_exit}个，"
            f"提报失败复提池{len(resubmit_pool['failed_resubmit_skus'])}个，"
            f"成功但需复提{len(resubmit_pool['success_resubmit_skus'])}个，"
            f"本次额外加入复提候选{added_resubmit}个，剩余{len(step6)}个。"
        )
        add_audit(audit, "提报状态过滤与复提合并", "通过", f"普通成功剔除 {removed_success}，额外加入复提 {added_resubmit}")
    else:
        current_success_file = str(config.get("current_success_skus_file", "") or "").strip()
        submitted_file = str(config.get("submitted_skus_file", "") or "").strip()
        step6 = step5
        removed_success = 0
        removed_submitted = 0
        if current_success_file:
            step6, removed_success = remove_skus(step6, sku_col, load_sku_set(current_success_file, sku_col))
            add_audit(audit, "旧名单兼容：成功SKU", "通过", f"已读取并剔除 {removed_success} 个")
        if submitted_file:
            step6, removed_submitted = remove_skus(step6, sku_col, load_sku_set(submitted_file, sku_col))
            add_audit(audit, "旧名单兼容：已提报SKU", "通过", f"已读取并剔除 {removed_submitted} 个")
        if current_success_file or submitted_file:
            logs.append(f"6、未检测到最终结果中的提报状态字段，已按旧名单兼容逻辑处理：删除成功{removed_success}个、已提报{removed_submitted}个，剩余{len(step6)}个。")
        else:
            logs.append(f"6、未检测到最终结果中的提报状态字段（提报状态_最终/提报结果标签/提报详情），按初次提报处理，不做提报状态剔除，剩余{len(step6)}个。")
        add_audit(audit, "提报状态字段", "通过", "未检测到提报状态_最终/提报结果标签/提报详情；按初次提报处理，不需要提供旧成功/已提报名单")

    excluded_erps = set(config.get("exclude_erp_accounts", []))
    removed_erp = [row for row in step6 if str(row.get(erp_col, "")).strip() in excluded_erps]
    final_rows = [row for row in step6 if str(row.get(erp_col, "")).strip() not in excluded_erps]
    logs.append(f"7、剔除指定ERP帐号（{'、'.join(excluded_erps)}）{len(removed_erp)}个，剩余{len(final_rows)}个。")

    return write_filter_outputs(config, output_dir, date_tag, headers, final_rows, logs, audit, sku_col, source, len(source_rows), source_type)
