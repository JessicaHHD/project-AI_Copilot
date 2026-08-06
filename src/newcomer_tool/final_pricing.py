from __future__ import annotations

from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
import shutil
from typing import Any
import warnings

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

from .batch import batch_name, final_pricing_output_path, load_manifest, safe_filename_part, unique_path
from .config import AppConfig


warnings.filterwarnings("ignore", message="Workbook contains no default style, apply openpyxl's default")


def normalize(value: str) -> str:
    return value.replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")


def find_column(headers: list[str], candidates: list[str], required: bool = True, excludes: list[str] | None = None) -> str | None:
    excludes = excludes or []
    normalized = {normalize(header): header for header in headers}
    for candidate in candidates:
        compact_candidate = normalize(candidate)
        if compact_candidate in normalized:
            return normalized[compact_candidate]
    for header in headers:
        compact = normalize(header)
        if any(exclude and normalize(exclude) in compact for exclude in excludes):
            continue
        if any(normalize(candidate) in compact for candidate in candidates if candidate):
            return header
    if required:
        raise KeyError("缺少必要字段：" + "、".join([item for item in candidates if item]))
    return None


def to_decimal(value: object) -> Decimal | None:
    text = str(value or "").strip().replace(",", "")
    if not text or text in {"-", "--"}:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def money(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)



def format_value(value: object) -> object:
    if isinstance(value, Decimal):
        return float(money(value))
    return value



def read_sheet(path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheets = [workbook[sheet_name]] if sheet_name and sheet_name in workbook.sheetnames else list(workbook.worksheets)
        for worksheet in worksheets:
            headers, rows = read_worksheet(worksheet)
            if any(headers):
                return headers, rows
        return [], []
    finally:
        workbook.close()


def workbook_has_sheet(path: Path, sheet_name: str) -> bool:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return sheet_name in workbook.sheetnames
    finally:
        workbook.close()


def read_filter_logs(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if "筛品日志" not in workbook.sheetnames:
            return ["未读取到筛品日志：源筛品结果文件无“筛品日志”sheet。"]
        worksheet = workbook["筛品日志"]
        if hasattr(worksheet, "reset_dimensions"):
            worksheet.reset_dimensions()
        lines: list[str] = []
        for row in worksheet.iter_rows(values_only=True):
            values = [str(value).strip() for value in row if str(value or "").strip()]
            if values:
                lines.append(" | ".join(values))
        return lines or ["筛品日志为空。"]
    finally:
        workbook.close()


def read_worksheet(worksheet) -> tuple[list[str], list[dict[str, Any]]]:
    if hasattr(worksheet, "reset_dimensions"):
        worksheet.reset_dimensions()
    rows_iter = worksheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(rows_iter)]
    except StopIteration:
        return [], []
    rows: list[dict[str, Any]] = []
    for row in rows_iter:
        rows.append({headers[index]: value for index, value in enumerate(row[: len(headers)])})
    return headers, rows


def latest_filter_file(config: AppConfig) -> Path:
    configured = str(config.get("final_filter_file", "") or "").strip()
    if configured:
        path = Path(configured)
        if path.exists():
            return path
        raise FileNotFoundError(f"筛品结果文件不存在：{path}")
    candidates = sorted(
        config.filter_output_dir.glob("新人价筛品结果_*.xlsx"),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        raise FileNotFoundError(f"未找到筛品结果文件：{config.filter_output_dir / '新人价筛品结果_*.xlsx'}")
    return candidates[0]


def configured_price_directory(config: AppConfig) -> Path:
    return Path(str(config.get("final_price_dir", "") or config.download_output_dir))


def directory_price_files(config: AppConfig) -> list[Path]:
    directory = Path(str(config.get("final_price_dir", "") or config.download_output_dir))
    pattern = str(config.get("final_price_pattern", "*.xlsx") or "*.xlsx")
    if not directory.exists():
        raise FileNotFoundError(f"查价结果目录不存在：{directory}")
    files = sorted(path for path in directory.glob(pattern) if not path.name.startswith("~$"))
    if not files:
        raise FileNotFoundError(f"未找到查价结果文件：{directory / pattern}")
    return files


def manifest_price_files(config: AppConfig) -> tuple[list[Path], list[str]]:
    manifest = load_manifest(config)
    step = manifest.get("steps", {}).get("阶段1-Outlook查价下载", {})
    raw_files = step.get("查价结果文件") or []
    notes: list[str] = []
    if isinstance(raw_files, str):
        raw_files = [raw_files]
    files = [Path(str(path)) for path in raw_files if str(path or "").strip()]
    if not files:
        notes.append("批次清单中没有记录本次查价结果文件，改用目录扫描。")
        return [], notes
    missing = [path for path in files if not path.exists()]
    if missing:
        notes.append("批次清单中的部分查价结果文件不存在，改用目录扫描：" + "；".join(str(path) for path in missing))
        return [], notes
    return files, notes


def price_header_candidates(config: AppConfig) -> tuple[list[str], list[str]]:
    sku_candidates = [str(config.get("final_price_sku_column", "skuId")), "skuId", "skuid", "sku", "SKU", "SKU(含影)"]
    price_candidates = [
        str(config.get("final_price_value_column", "100天最低价")),
        "近N天最低价",
        "N天最低价",
        "100天最低价",
        "100天最低价格",
        "最低价",
        "价格",
        "100天最低",
    ]
    return sku_candidates, price_candidates


def obvious_registration_status_headers(headers: list[str]) -> bool:
    text = " ".join(str(header or "") for header in headers)
    keywords = ["提报状态", "提报结果", "促销状态", "报名状态", "审核状态", "提报详情", "失败原因", "促销ID", "促销名称"]
    return sum(1 for keyword in keywords if keyword in text) >= 2


def validate_price_file(path: Path, config: AppConfig) -> tuple[bool, str]:
    try:
        headers, _ = read_sheet(path, None)
    except Exception as exc:
        return False, f"文件无法读取：{exc}"
    if not headers:
        return False, "文件没有可识别表头"
    sku_candidates, price_candidates = price_header_candidates(config)
    sku_col = find_column(headers, sku_candidates, required=False)
    status_like = obvious_registration_status_headers(headers)
    price_excludes = ["促销名称", "促销ID", "促销创建人", "创建人", "ID"]
    explicit_price_col = find_column(headers, [candidate for candidate in price_candidates if candidate != "价格"], required=False)
    price_col = explicit_price_col or (None if status_like else find_column(headers, price_candidates, required=False, excludes=price_excludes))
    if not sku_col:
        return False, "未识别到SKU字段"
    if not price_col:
        if status_like:
            return False, "疑似提报/审核情况导出文件，不是查价结果"
        return False, "未识别到查价价格字段"
    return True, ""


def quarantine_suspicious_file(path: Path, reason: str) -> Path:
    target_dir = path.parent / "疑似误下载"
    target_dir.mkdir(parents=True, exist_ok=True)
    target = unique_path(target_dir / path.name)
    try:
        shutil.move(str(path), str(target))
    except OSError as exc:
        raise RuntimeError(f"疑似误下载文件隔离失败：{path}。原因：{reason}。请关闭相关 Excel 文件后重试。") from exc
    return target


def select_price_files(config: AppConfig) -> dict[str, Any]:
    notes: list[str] = []
    files, manifest_notes = manifest_price_files(config)
    notes.extend(manifest_notes)
    source = "批次清单"
    if not files:
        source = "目录回退"
        files = directory_price_files(config)

    valid_files: list[Path] = []
    quarantined: list[dict[str, str]] = []
    for file in files:
        if file.name.startswith("~$"):
            continue
        ok, reason = validate_price_file(file, config)
        if ok:
            valid_files.append(file)
            continue
        target = quarantine_suspicious_file(file, reason)
        quarantined.append({"source": str(file), "target": str(target), "reason": reason})

    if not valid_files:
        detail = "；".join(item["reason"] + "：" + item["source"] for item in quarantined)
        raise FileNotFoundError(f"未找到可用于整合的查价结果文件。{detail or '请确认本次Outlook下载的是查价结果。'}")

    return {
        "files": valid_files,
        "source": source,
        "candidate_count": len(files),
        "quarantined": quarantined,
        "notes": notes,
    }


def price_files(config: AppConfig) -> list[Path]:
    return list(select_price_files(config)["files"])


def latest_resubmit_reference_file(config: AppConfig) -> Path | None:
    configured = str(config.get("resubmit_final_file", "") or "").strip()
    if configured:
        path = Path(configured)
        if not path.exists():
            raise FileNotFoundError(f"复提最终结果文件不存在：{path}")
        return path
    directory = config.output_root / "最终结果"
    if not directory.exists():
        return None
    patterns = ["*提报情况整理表*.xlsx", "新人价提报整理_*.xlsx"]
    candidates: list[Path] = []
    for pattern in patterns:
        candidates.extend(path for path in directory.glob(pattern) if path.is_file() and not path.name.startswith("~$"))
    candidates = sorted(set(candidates), key=lambda item: item.stat().st_mtime, reverse=True)
    return candidates[0] if candidates else None


def date_label(config: AppConfig) -> str:
    configured = str(config.get("final_price_date_label", "") or "").strip()
    if configured:
        return configured
    now = datetime.now()
    return f"{now.month}.{now.day}"


def writable_output_file(path: Path) -> Path:
    if not path.exists():
        return path
    try:
        with path.open("a+b"):
            return path
    except PermissionError:
        stamp = datetime.now().strftime("%H%M%S")
        return path.with_name(f"{path.stem}_{stamp}{path.suffix}")


MANAGED_SHEETS = ["筛选结果", "价格复核剔除", "价格复核明细", "查价结果合并", "ERP汇总", "整合日志"]


def current_final_pricing_file(config: AppConfig, output_dir: Path) -> Path | None:
    manifest = load_manifest(config)
    step = manifest.get("steps", {}).get("阶段1-筛品查价整合", {})
    manifest_file = str(step.get("筛品查价表") or "").strip()
    if manifest_file:
        path = Path(manifest_file)
        if path.exists() and path.is_file() and not path.name.startswith("~$"):
            return path

    pattern = f"{batch_name(config)}_筛品查价表*.xlsx"
    candidates = sorted(
        (
            path
            for path in output_dir.glob(pattern)
            if path.is_file()
            and not path.name.startswith("~$")
            and "提报情况整理表" not in path.name
            and "业务确认" not in path.name
        ),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def ensure_existing_workbook_writable(path: Path) -> None:
    try:
        with path.open("a+b"):
            return
    except PermissionError as exc:
        raise RuntimeError(f"当前筛品查价表正在被占用，无法原文件更新。请关闭 Excel 后重试：{path}") from exc


def backup_existing_workbook(path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_dir = path.parent / "_archive" / f"{timestamp}_final_pricing_update"
    archive_dir.mkdir(parents=True, exist_ok=True)
    target = unique_path(archive_dir / path.name)
    try:
        shutil.copy2(path, target)
    except OSError as exc:
        raise RuntimeError(f"更新前备份筛品查价表失败：{path}。请关闭相关 Excel 文件后重试。") from exc
    return target


def load_existing_main_rows(path: Path | None, sku_column_name: str) -> tuple[list[str], dict[str, dict[str, Any]], list[dict[str, Any]]]:
    if path is None or not path.exists():
        return [], {}, []
    if not workbook_has_sheet(path, "筛选结果"):
        return [], {}, []
    headers, rows = read_sheet(path, "筛选结果")
    if not headers:
        return [], {}, []
    sku_col = find_column(headers, [sku_column_name, "SKU(含影)", "skuid", "skuId", "SKU", "sku"], required=False)
    rows_by_sku: dict[str, dict[str, Any]] = {}
    if sku_col:
        for row in rows:
            sku = str(row.get(sku_col, "") or "").strip()
            if sku:
                rows_by_sku[sku] = dict(row)
    return headers, rows_by_sku, rows


def load_existing_price_review_rows(path: Path | None) -> list[dict[str, Any]]:
    if path is None or not path.exists():
        return []
    if not workbook_has_sheet(path, "价格复核明细"):
        return []
    headers, rows = read_sheet(path, "价格复核明细")
    if not headers:
        return []
    return [dict(row) for row in rows]


def merge_rows_preserving_existing(existing_row: dict[str, Any] | None, current_row: dict[str, Any]) -> dict[str, Any]:
    merged = dict(existing_row or {})
    for key, value in current_row.items():
        if value is None or str(value).strip() == "":
            merged.setdefault(key, value)
        else:
            merged[key] = value
    return merged


def combine_price_review_rows(existing_rows: list[dict[str, Any]], current_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    combined: dict[tuple[str, str, str], dict[str, Any]] = {}
    for row in [*existing_rows, *current_rows]:
        sku = str(row.get("SKU", "") or "").strip()
        previous_col = str(row.get("上次查价字段", "") or "").strip()
        current_col = str(row.get("本次查价字段", "") or "").strip()
        key = (sku, previous_col, current_col)
        combined[key] = dict(row)
    return list(combined.values())


def remove_managed_sheets(workbook: Workbook) -> None:
    for sheet_name in MANAGED_SHEETS:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]


def merge_price_rows(files: list[Path], config: AppConfig) -> tuple[list[str], list[dict[str, Any]], dict[str, Decimal], list[str]]:
    all_headers: list[str] = []
    merged_rows: list[dict[str, Any]] = []
    price_by_sku: dict[str, Decimal] = {}
    duplicate_skus: Counter[str] = Counter()
    sku_candidates, price_candidates = price_header_candidates(config)
    price_excludes = ["促销名称", "促销ID", "促销创建人", "创建人", "ID"]

    for file in files:
        headers, rows = read_sheet(file, None)
        if not headers:
            continue
        for header in headers:
            if header and header not in all_headers:
                all_headers.append(header)
        try:
            sku_col = find_column(headers, sku_candidates)
            price_col = find_column(headers, price_candidates, excludes=price_excludes)
        except KeyError as exc:
            raise KeyError(f"{file.name}：{exc}") from exc
        for row in rows:
            row = dict(row)
            row["来源文件"] = file.name
            merged_rows.append(row)
            sku = str(row.get(sku_col, "") or "").strip()
            price = to_decimal(row.get(price_col))
            if not sku or price is None:
                continue
            if sku in price_by_sku:
                duplicate_skus[sku] += 1
                price_by_sku[sku] = min(price_by_sku[sku], price)
            else:
                price_by_sku[sku] = price

    if "来源文件" not in all_headers:
        all_headers.append("来源文件")
    logs = [
        f"查价文件数：{len(files)}",
        f"查价明细行数：{len(merged_rows)}",
        f"可匹配价格SKU数：{len(price_by_sku)}",
        f"重复SKU数：{len(duplicate_skus)}，重复时取100天最低价较小值。",
    ]
    return all_headers, merged_rows, price_by_sku, logs


def calculate_newcomer_price(base_price: Decimal | None, mapping: str) -> Decimal | None:
    if base_price is None:
        return None
    subsidy = Decimal("8") if mapping.strip() == "自营" else Decimal("6.5")
    result = base_price - subsidy
    if result <= 0:
        return Decimal("0.01")
    return money(result)


def config_decimal(config: AppConfig, key: str, default: str) -> Decimal:
    value = config.get(key, default)
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return Decimal(default)


def config_list(config: AppConfig, key: str) -> list[str]:
    value = config.get(key, [])
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value or "").strip()
    return [text] if text else []


def contains_any_keyword(text: str, keywords: list[str]) -> bool:
    return any(keyword and keyword in text for keyword in keywords)


def is_resubmit_candidate(row: dict[str, Any], status_col: str | None, label_col: str | None, detail_col: str | None, config: AppConfig) -> bool:
    if not status_col:
        return False
    status = str(row.get(status_col, "") or "")
    if "提报失败" in status:
        return True
    if "提报成功" not in status:
        return False
    label = str(row.get(label_col, "") or "") if label_col else ""
    detail = str(row.get(detail_col, "") or "") if detail_col else ""
    return contains_any_keyword(f"{label} {detail}", config_list(config, "resubmit_success_keywords"))


def price_review_column_candidates(headers: list[str], current_lookup_column: str) -> list[str]:
    excluded_fragments = ["新人价", "差值", "差异", "比例", "状态", "结果", "说明", "文件", "匹配", "复核", "提报"]
    candidates: list[str] = []
    for header in headers:
        text = str(header or "").strip()
        if not text or text == current_lookup_column or "查价" not in text:
            continue
        if any(fragment in text for fragment in excluded_fragments):
            continue
        candidates.append(text)
    return candidates


def find_previous_price(row: dict[str, Any], price_columns: list[str]) -> tuple[str, Decimal | None]:
    fallback_column = price_columns[-1] if price_columns else ""
    for column in reversed(price_columns):
        price = to_decimal(row.get(column))
        if price is not None:
            return column, price
    return fallback_column, None


def load_resubmit_reference(config: AppConfig, current_lookup_column: str, sku_column_name: str) -> tuple[Path | None, dict[str, dict[str, Any]], list[str], list[str]]:
    path = latest_resubmit_reference_file(config)
    if path is None:
        return None, {}, [], []
    headers, rows = read_sheet(path, "筛选结果")
    if not headers:
        return path, {}, [], []
    sku_col = find_column(headers, [sku_column_name, "SKU(含影)", "skuid", "skuId", "SKU", "sku"], required=False)
    status_col = find_column(headers, ["提报状态_最终", "提报状态"], required=False)
    label_col = find_column(headers, ["提报结果标签", "提报失败标签", "促销状态"], required=False)
    detail_col = find_column(headers, ["提报详情", "详情", "失败原因"], required=False)
    if not sku_col or not status_col:
        return path, {}, headers, []
    rows_by_sku: dict[str, dict[str, Any]] = {}
    for row in rows:
        sku = str(row.get(sku_col, "") or "").strip()
        if not sku or not is_resubmit_candidate(row, status_col, label_col, detail_col, config):
            continue
        rows_by_sku[sku] = dict(row)
    return path, rows_by_sku, headers, price_review_column_candidates(headers, current_lookup_column)


def review_rate_text(rate: Decimal | None) -> object:
    if rate is None:
        return ""
    return float((rate * Decimal("100")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def build_price_review_row(
    sku: str,
    product_name: str,
    previous_column: str,
    previous_price: Decimal | None,
    current_column: str,
    current_price: Decimal | None,
    abs_threshold: Decimal,
    rate_threshold: Decimal,
) -> tuple[dict[str, Any], bool]:
    diff: Decimal | None = None
    rate: Decimal | None = None
    result = "通过"
    note = "价格差异在阈值内"
    excluded = False

    if previous_price is None:
        result = "上次查价缺失"
        note = "未找到上一轮可用于对比的查价价格，默认不因价格差异剔除"
    elif current_price is None:
        result = "本次查价缺失"
        note = "本次查价未匹配，按现有未匹配查价逻辑处理"
    else:
        diff = money(current_price - previous_price)
        abs_diff = abs(diff)
        if previous_price != 0:
            rate = abs_diff / abs(previous_price)
        reasons: list[str] = []
        if abs_diff > abs_threshold:
            reasons.append(f"价格差值绝对值超过{money(abs_threshold)}元")
        if rate is not None and rate > rate_threshold:
            reasons.append(f"价格差异比例超过{review_rate_text(rate_threshold)}%")
        if reasons:
            result = "剔除"
            note = "；".join(reasons)
            excluded = True

    row = {
        "SKU": sku,
        "商品名称": product_name,
        "上次查价字段": previous_column,
        "上次查价": money(previous_price) if previous_price is not None else "",
        "本次查价字段": current_column,
        "本次查价": money(current_price) if current_price is not None else "",
        "价格差值": diff if diff is not None else "",
        "价格差异比例": review_rate_text(rate),
        "价格复核结果": result,
        "说明": note,
        "剔除原因": note if excluded else "",
    }
    return row, excluded


def write_rows(workbook: Workbook, title: str, headers: list[str], rows: list[dict[str, Any]], index: int | None = None) -> None:
    sheet = workbook.create_sheet(title, index=index)
    sheet.append(headers)
    for row in rows:
        sheet.append([format_value(row.get(header, "")) for header in headers])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(column[0].value or "")) + 2, 12), 32)


def write_log_sheet(workbook: Workbook, logs: list[str]) -> None:
    sheet = workbook.create_sheet("整合日志")
    for line in logs:
        sheet.append([line])
    sheet.column_dimensions["A"].width = 120
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def unique_erp_values(rows: list[dict[str, Any]], erp_column: str) -> list[str]:
    values = {
        str(row.get(erp_column, "") or "").strip()
        for row in rows
        if str(row.get(erp_column, "") or "").strip()
    }
    return sorted(values)


def write_erp_sheet(workbook: Workbook, erp_values: list[str]) -> str:
    joined_text = ";".join(erp_values)
    sheet = workbook.create_sheet("ERP汇总")
    sheet.append(["销售员ERP去重合并", joined_text])
    sheet.append([])
    sheet.append(["销售员ERP帐号"])
    for value in erp_values:
        sheet.append([value])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet[3][0].font = Font(bold=True)
    sheet.freeze_panes = "A4"
    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 120
    sheet[1][1].alignment = Alignment(wrap_text=True, vertical="top")
    return joined_text


def run_final_pricing(config: AppConfig) -> dict[str, Any]:
    filter_file = latest_filter_file(config)
    price_selection = select_price_files(config)
    files = list(price_selection["files"])
    label = date_label(config)
    lookup_column = f"{label}_查价"
    output_dir = config.output_root / "最终结果"
    output_dir.mkdir(parents=True, exist_ok=True)
    existing_output_file = current_final_pricing_file(config, output_dir)

    filter_headers, filter_rows = read_sheet(filter_file, "筛选结果")
    if not filter_headers:
        raise ValueError(f"筛品结果文件为空：{filter_file}")
    filter_logs = read_filter_logs(filter_file)
    sku_col = find_column(filter_headers, [str(config.get("sku_column")), "SKU(含影)", "sku", "SKU"])
    mapping_col = find_column(filter_headers, [str(config.get("mapping_column")), "1自营3POP_映射"])
    erp_col = find_column(filter_headers, [str(config.get("erp_column")), "销售员ERP帐号", "销售员ERP账号"])
    product_name_col = find_column(filter_headers, [str(config.get("product_name_column")), "sku名称", "商品名称", "商品名称（含影）"], required=False)
    existing_headers, existing_rows_by_sku, _ = load_existing_main_rows(existing_output_file, sku_col)
    existing_price_review_rows = load_existing_price_review_rows(existing_output_file)

    price_headers, price_rows, price_by_sku, price_logs = merge_price_rows(files, config)
    price_review_enabled = bool(config.get("resubmit_price_review_enabled", True))
    price_review_detail_enabled = bool(config.get("resubmit_price_review_detail_sheet", True))
    price_review_abs_threshold = config_decimal(config, "resubmit_price_diff_abs_threshold", "10")
    price_review_rate_threshold = config_decimal(config, "resubmit_price_diff_rate_threshold", "0.2")
    resubmit_reference_file: Path | None = None
    resubmit_reference_rows: dict[str, dict[str, Any]] = {}
    previous_price_columns: list[str] = []
    if price_review_enabled:
        resubmit_reference_file, resubmit_reference_rows, _, previous_price_columns = load_resubmit_reference(config, lookup_column, sku_col)
    final_headers = list(existing_headers or filter_headers)
    for header in filter_headers:
        if header not in final_headers:
            final_headers.append(header)
    for header in [lookup_column, "新人价"]:
        if header not in final_headers:
            final_headers.append(header)

    matched_count = 0
    missing_count = 0
    removed_count = 0
    price_review_passed_count = 0
    price_review_excluded_count = 0
    previous_price_missing_count = 0
    current_price_missing_review_count = 0
    price_review_rows: list[dict[str, Any]] = []
    price_review_excluded_rows: list[dict[str, Any]] = []
    final_rows: list[dict[str, Any]] = []
    for row in filter_rows:
        current_row = dict(row)
        sku = str(current_row.get(sku_col, "") or "").strip()
        row = merge_rows_preserving_existing(existing_rows_by_sku.get(sku), current_row)
        base_price = price_by_sku.get(sku)
        newcomer_price = calculate_newcomer_price(base_price, str(row.get(mapping_col, "") or ""))
        reference_row = resubmit_reference_rows.get(sku) if price_review_enabled else None
        review_row: dict[str, Any] | None = None
        review_excluded = False
        if reference_row:
            product_name = str(row.get(product_name_col, "") or reference_row.get(product_name_col or "", "") or "") if product_name_col else ""
            previous_column, previous_price = find_previous_price(reference_row, previous_price_columns)
            review_row, review_excluded = build_price_review_row(
                sku,
                product_name,
                previous_column,
                previous_price,
                lookup_column,
                base_price,
                price_review_abs_threshold,
                price_review_rate_threshold,
            )
            price_review_rows.append(review_row)
            if previous_price is None:
                previous_price_missing_count += 1
            if base_price is None:
                current_price_missing_review_count += 1
        if base_price is None:
            missing_count += 1
            row[lookup_column] = ""
            row["新人价"] = ""
            final_rows.append(row)
            continue
        matched_count += 1
        row[lookup_column] = money(base_price)
        row["新人价"] = newcomer_price
        if review_row:
            if review_excluded:
                price_review_excluded_count += 1
                price_review_excluded_rows.append(review_row)
                continue
            price_review_passed_count += 1
        if newcomer_price is not None and newcomer_price >= Decimal("200"):
            removed_count += 1
            continue
        final_rows.append(row)

    updated_existing = existing_output_file is not None
    output_file = existing_output_file or final_pricing_output_path(config, output_dir, label)
    backup_file: Path | None = None
    if updated_existing:
        ensure_existing_workbook_writable(output_file)
        backup_file = backup_existing_workbook(output_file)

    combined_price_review_rows = combine_price_review_rows(existing_price_review_rows, price_review_rows)

    logs = [
        "【筛品日志】",
        *filter_logs,
        "",
        "【最终整合日志】",
        f"筛品结果文件：{filter_file}",
        f"查价日期字段：{lookup_column}",
        f"查价文件来源：{price_selection['source']}",
        f"查价候选文件数：{price_selection['candidate_count']}",
        f"疑似误下载隔离数：{len(price_selection['quarantined'])}",
        *[f"疑似误下载隔离：{item['source']} -> {item['target']}（{item['reason']}）" for item in price_selection["quarantined"]],
        *[f"查价文件选择提示：{note}" for note in price_selection["notes"]],
        f"筛品查价表写入方式：{'更新已有表' if updated_existing else '新建表'}",
        f"更新前备份文件：{backup_file or '无'}",
        f"历史价格复核明细保留行数：{len(existing_price_review_rows)}",
        *price_logs,
        f"筛品结果行数：{len(filter_rows)}",
        f"查价匹配行数：{matched_count}",
        f"未匹配查价行数：{missing_count}",
        f"复提价格复核：{'启用' if price_review_enabled else '未启用'}",
        f"复提价格复核来源表：{resubmit_reference_file or '未找到'}",
        f"复提价格复核上次查价候选字段：{'、'.join(previous_price_columns) if previous_price_columns else '未找到'}",
        f"复提价格差异绝对值阈值：{money(price_review_abs_threshold)}",
        f"复提价格差异比例阈值：{review_rate_text(price_review_rate_threshold)}%",
        f"参与复提价格复核SKU数：{len(price_review_rows)}",
        f"价格复核通过数：{price_review_passed_count}",
        f"价格复核剔除数：{price_review_excluded_count}",
        f"上次查价缺失数：{previous_price_missing_count}",
        f"本次查价缺失数：{current_price_missing_review_count}",
        f"剔除新人价大于等于200元：{removed_count}",
        f"最终保留行数：{len(final_rows)}",
    ]
    erp_values = unique_erp_values(final_rows, erp_col)
    logs.append(f"最终保留商品去重销售员ERP数：{len(erp_values)}")

    if updated_existing:
        workbook = load_workbook(output_file)
        remove_managed_sheets(workbook)
    else:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)

    write_rows(workbook, "筛选结果", final_headers, final_rows, index=0)
    if price_review_excluded_rows:
        write_rows(
            workbook,
            "价格复核剔除",
            ["SKU", "商品名称", "上次查价字段", "上次查价", "本次查价字段", "本次查价", "价格差值", "价格差异比例", "剔除原因"],
            price_review_excluded_rows,
        )
    if price_review_detail_enabled and combined_price_review_rows:
        write_rows(
            workbook,
            "价格复核明细",
            ["SKU", "商品名称", "上次查价字段", "上次查价", "本次查价字段", "本次查价", "价格差值", "价格差异比例", "价格复核结果", "说明"],
            combined_price_review_rows,
        )
    write_rows(workbook, "查价结果合并", price_headers, price_rows)
    erp_joined_text = write_erp_sheet(workbook, erp_values)
    write_log_sheet(workbook, logs)
    workbook.save(output_file)

    return {
        "filter_file": filter_file,
        "price_files": files,
        "price_file_source": price_selection["source"],
        "price_file_candidate_count": price_selection["candidate_count"],
        "misdownload_files": price_selection["quarantined"],
        "misdownload_count": len(price_selection["quarantined"]),
        "output_file": output_file,
        "updated_existing": updated_existing,
        "backup_file": backup_file,
        "lookup_column": lookup_column,
        "source_rows": len(filter_rows),
        "matched_rows": matched_count,
        "missing_rows": missing_count,
        "removed_rows": removed_count,
        "price_review_rows": len(price_review_rows),
        "price_review_detail_rows": len(combined_price_review_rows),
        "price_review_excluded_rows": price_review_excluded_count,
        "final_rows": len(final_rows),
        "erp_count": len(erp_values),
        "erp_joined_text": erp_joined_text,
        "logs": logs,
    }
